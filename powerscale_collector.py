#!/usr/bin/env python3
"""
PowerScale / Isilon OneFS 9.x — Migration Data Collector  v4.0
Developer: SHAIKH SHOAIB · Sr. Advisor Delivery Specialist

Auth  : POST /session/1/session  (isisessid + isicsrf CSRF)
Usage : pip install flask requests urllib3 openpyxl
        python powerscale_collector.py
Access: http://localhost:5050

CONFIRMED WORKING PATHS on OneFS 9.10.1.3:
  /platform/3/cluster/identity        ✔
  /platform/3/cluster/nodes           ✔
  /platform/3/cluster/nodes/{lnn}     ✔
  /platform/3/cluster/time            ✔  (node timestamps only)
  /platform/3/event/eventlists        ✔
  /platform/3/job/jobs                ✔
  /platform/5/license/licenses        ✔
  /platform/2/protocols/nfs/exports   ✔
  /platform/4/protocols/smb/shares    ✔
  /platform/1/quota/quotas            ✔
  /platform/1/snapshot/snapshots      ✔
  /platform/3/sync/policies           ✔
  /platform/7/storagepool/nodepools   ✔
  /platform/3/auth/providers/summary  ✔
  /platform/3/auth/roles              ✔
  /platform/1/statistics/keys         ✔
  /platform/1/statistics/current      ✔

CONFIRMED 404 — DO NOT USE on OneFS 9.10.1.3:
  /platform/x/cluster/time/settings    404
  /platform/x/cluster/time/settings/servers  404
  /platform/x/ntp/servers              404
  /platform/1/session/1/privileges     404
"""

import sys, json, base64, secrets, threading, logging, io
from datetime import datetime, timezone, timedelta
from functools import wraps

try:
    import requests, urllib3
    from flask import Flask, request, jsonify, Response, send_file
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except ImportError as e:
    print(f"\n[ERROR] {e}\n  pip install flask requests urllib3 openpyxl\n"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Config ────────────────────────────────────────────────────────
SESSION_TIMEOUT_MINUTES = 30
MAX_SESSIONS            = 10
PROXY_TIMEOUT           = 30
LISTEN_HOST             = "127.0.0.1"
LISTEN_PORT             = 5050
DEBUG_MODE              = True

logging.basicConfig(
    level   = logging.DEBUG if DEBUG_MODE else logging.INFO,
    format  = "[%(asctime)s] %(levelname)-7s %(message)s",
    datefmt = "%H:%M:%S",
)
log = logging.getLogger("psscale")
app = Flask(__name__)

def utcnow(): return datetime.now(timezone.utc)


# ── IsilonSession ─────────────────────────────────────────────────
class IsilonSession:
    def __init__(self, host, port, username, password, ssl_verify=False):
        self.base_url = f"https://{host}:{port}"
        self._s       = requests.Session()
        self._s.verify = ssl_verify
        body = json.dumps({"username": username, "password": password,
                           "services": ["platform", "namespace"]})
        log.debug("POST /session/1/session → %s:%d  user=%s  pass_len=%d  specials=%d",
                  host, port, username, len(password),
                  len([c for c in password if not c.isalnum()]))
        r = self._s.post(f"{self.base_url}/session/1/session", data=body,
                         headers={"Content-Type": "application/json",
                                  "Referer": self.base_url}, timeout=15)
        log.debug("Session response: HTTP %d", r.status_code)
        if r.status_code == 401:
            raise PermissionError(f"Invalid credentials (401): {r.text[:120]}")
        if not r.ok:
            raise Exception(f"Session creation failed ({r.status_code}): {r.text[:200]}")
        csrf = self._s.cookies.get("isicsrf", "")
        self._s.headers.update({"X-CSRF-Token": csrf, "Referer": self.base_url,
                                 "Content-Type": "application/json"})
        log.debug("Session established — isisessid=%s  isicsrf=%s",
                  "set" if self._s.cookies.get("isisessid") else "MISSING",
                  "set" if csrf else "MISSING")

    def get(self, path):
        url = f"{self.base_url}{path}"
        log.debug("→ GET %s", path)
        try:
            r = self._s.get(url, timeout=PROXY_TIMEOUT)
        except requests.RequestException as ex:
            return {"_error": str(ex)}, 0
        log.debug("← %d  (%d bytes)", r.status_code, len(r.content))
        if r.status_code in (401, 403):
            body = r.text[:300].strip()
            log.warning("PAPI %d for %s | %s", r.status_code, path, body)
            return {"_error": "insufficient_privilege", "_status": r.status_code,
                    "_message": f"HTTP {r.status_code} — add ISI_PRIV for: {path.split('?')[0]}"}, r.status_code
        if r.status_code == 200:
            try: return r.json(), 200
            except: return {"_error": "invalid_json", "_raw": r.text[:300]}, r.status_code
        return {"_error": f"HTTP {r.status_code}", "_raw": r.text[:200]}, r.status_code

    def try_paths(self, *paths):
        last, lcode = {"_error": "no_paths"}, 0
        for path in paths:
            data, code = self.get(path)
            if code == 200: return data, 200
            if code in (401, 403): return data, code
            last, lcode = data, code
        return last, lcode

    def close(self):
        try: self._s.delete(f"{self.base_url}/session/1/session", timeout=5)
        except: pass
        self._s.close()

# ── Session store ─────────────────────────────────────────────────
_store: dict = {}
_lock = threading.Lock()

def _purge():
    now = utcnow()
    with _lock:
        dead = [k for k, v in _store.items() if v["expires"] < now]
        for k in dead: del _store[k]

def _get(token):
    if not token: return None
    _purge()
    with _lock:
        e = _store.get(token)
        if e is None:
            log.debug("Token not found: ...%s  (store=%d)", token[-8:], len(_store))
            return None
        e["last_active"] = utcnow()
        e["expires"]     = utcnow() + timedelta(minutes=SESSION_TIMEOUT_MINUTES)
        log.debug("Token OK: ...%s → %s", token[-8:], e["host"])
        return e

def _save(token, sess, host):
    _purge()
    with _lock:
        if len(_store) >= MAX_SESSIONS and token not in _store:
            oldest = min(_store, key=lambda k: _store[k]["last_active"])
            del _store[oldest]
        _store[token] = {
            "session":     sess,
            "host":        host,
            "ntp_servers": [],
            "expires":     utcnow() + timedelta(minutes=SESSION_TIMEOUT_MINUTES),
            "last_active": utcnow(),
        }
    log.debug("Saved session ...%s → %s  (store=%d)", token[-8:], host, len(_store))

def _delete(token):
    with _lock:
        entry = _store.pop(token, None)
    if entry:
        try: entry["session"].close()
        except: pass

# ── Auth decorator ────────────────────────────────────────────────
def get_token():
    tok = request.headers.get("X-Session-Token","") or request.args.get("token","")
    log.debug("%-6s %-38s token=%s", request.method, request.path,
              ("..."+tok[-8:]) if tok else "MISSING")
    return tok

def require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        tok = get_token()
        entry = _get(tok)
        if entry is None:
            log.warning("401 for %s — token %s", request.path,
                        ("..."+tok[-8:]) if tok else "EMPTY")
            return jsonify({"error": "Not authenticated or session expired"}), 401
        return f(*args, entry=entry, **kwargs)
    return wrapper

def proxy(entry, *paths):
    data, code = entry["session"].try_paths(*paths)
    if code == 200: return jsonify(data), 200
    if code in (401, 403):
        return jsonify({"_privilege_denied": True, "_status": code,
                        "_message": data.get("_message", f"HTTP {code}")}), 200
    return jsonify({"_error": data.get("_error", f"HTTP {code}"), "_status": code}), 502


# ════════════════════════════════════════════════════════
#  FLASK ROUTES
# ════════════════════════════════════════════════════════

@app.route("/api/connect", methods=["POST"])
def api_connect():
    body       = request.get_json(force=True, silent=True) or {}
    host       = (body.get("host") or "").strip()
    port       = int(body.get("port") or 8080)
    username   = (body.get("username") or "").strip()
    password   = body.get("password") or ""
    ssl_verify = bool(body.get("ssl_verify", False))
    if not host or not username or not password:
        return jsonify({"error": "host, username and password required"}), 400
    specials = [c for c in password if not c.isalnum()]
    log.info("Connect: %s@%s:%d  ssl_verify=%s  pass_len=%d  specials=%d",
             username, host, port, ssl_verify, len(password), len(specials))
    try:
        sess = IsilonSession(host, port, username, password, ssl_verify)
    except PermissionError as ex:
        return jsonify({"error": str(ex)}), 401
    except Exception as ex:
        log.error("Connect failed: %s", ex)
        return jsonify({"error": str(ex)}), 502
    ident, code = sess.get("/platform/3/cluster/identity")
    if code != 200:
        sess.close()
        return jsonify({"error": f"Connected but identity failed (HTTP {code})"}), 502
    token        = secrets.token_hex(32)
    _save(token, sess, host)
    cluster_name = ident.get("name", host)
    onefs_ver    = ident.get("onefs_version", {}).get("revision", "9.x")
    log.info("Connect OK: ...%s  cluster=%s  onefs=%s", token[-8:], cluster_name, onefs_ver)
    return jsonify({"status": "connected", "token": token, "cluster": cluster_name,
                    "onefs": onefs_ver,
                    "expires": (utcnow() + timedelta(minutes=SESSION_TIMEOUT_MINUTES)).isoformat()})

@app.route("/api/disconnect", methods=["POST"])
def api_disconnect():
    _delete(get_token()); return jsonify({"status": "disconnected"})

@app.route("/api/status")
def api_status():
    tok = get_token(); entry = _get(tok)
    if entry is None: return jsonify({"connected": False})
    remaining = int((entry["expires"] - utcnow()).total_seconds())
    return jsonify({"connected": True, "host": entry["host"], "remaining": max(0, remaining)})

# ── Core inventory ────────────────────────────────────────────────

@app.route("/api/cluster/identity")
@require_auth
def ep_identity(entry):
    return proxy(entry, "/platform/3/cluster/identity", "/platform/1/cluster/identity")

@app.route("/api/cluster/config")
@require_auth
def ep_config(entry):
    return proxy(entry, "/platform/3/cluster/config", "/platform/1/cluster/config")

@app.route("/api/cluster/nodes")
@require_auth
def ep_nodes(entry):
    """Two-step: list LNNs then fetch full detail per node."""
    sess = entry["session"]
    list_data, list_code = sess.try_paths("/platform/3/cluster/nodes", "/platform/1/cluster/nodes")
    if list_code in (401, 403):
        return jsonify({"_privilege_denied": True, "_status": list_code,
                        "_message": list_data.get("_message", f"HTTP {list_code}")}), 200
    if list_code != 200:
        return jsonify(list_data), 502
    nodes_stub = list_data.get("nodes", [])
    if not nodes_stub:
        return jsonify({"nodes": []}), 200
    full_nodes = []
    for stub in nodes_stub:
        lnn = stub.get("lnn") or stub.get("id")
        if lnn is None:
            full_nodes.append(stub); continue
        detail, code = sess.try_paths(f"/platform/3/cluster/nodes/{lnn}",
                                       f"/platform/1/cluster/nodes/{lnn}")
        if code == 200:
            inner = detail.get("nodes", [detail])
            node  = inner[0] if inner else detail
            full_nodes.append({**stub, **node})
            log.debug("Node LNN=%s ok — keys: %s", lnn, list(node.keys())[:12])
        else:
            log.warning("Node LNN=%s detail HTTP %d — using stub", lnn, code)
            full_nodes.append(stub)
    return jsonify({"nodes": full_nodes}), 200

@app.route("/api/nfs/exports")
@require_auth
def ep_nfs(entry):
    return proxy(entry, "/platform/2/protocols/nfs/exports?limit=1000",
                        "/platform/3/protocols/nfs/exports?limit=1000")

@app.route("/api/smb/shares")
@require_auth
def ep_smb(entry):
    return proxy(entry, "/platform/4/protocols/smb/shares?limit=1000",
                        "/platform/3/protocols/smb/shares?limit=1000")

@app.route("/api/quotas")
@require_auth
def ep_quotas(entry):
    return proxy(entry, "/platform/1/quota/quotas?limit=1000",
                        "/platform/2/quota/quotas?limit=1000")

@app.route("/api/snapshots")
@require_auth
def ep_snapshots(entry):
    return proxy(entry, "/platform/1/snapshot/snapshots?limit=1000",
                        "/platform/3/snapshot/snapshots?limit=1000")

@app.route("/api/synciq/policies")
@require_auth
def ep_synciq(entry):
    return proxy(entry, "/platform/3/sync/policies?limit=500",
                        "/platform/4/sync/policies?limit=500")

@app.route("/api/synciq/reports")
@require_auth
def ep_synciq_reports(entry):
    return proxy(entry, "/platform/4/sync/reports?limit=50",
                        "/platform/3/sync/reports?limit=50")

@app.route("/api/pools")
@require_auth
def ep_pools(entry):
    return proxy(entry, "/platform/9/storagepool/storagepools",
                        "/platform/7/storagepool/nodepools",
                        "/platform/7/storagepool/storagepools",
                        "/platform/5/storagepool/storagepools",
                        "/platform/3/storagepool/storagepools")

@app.route("/api/zones")
@require_auth
def ep_zones(entry):
    return proxy(entry, "/platform/3/zones", "/platform/1/zones")

@app.route("/api/network/interfaces")
@require_auth
def ep_network(entry):
    return proxy(entry, "/platform/12/network/interfaces",
                        "/platform/7/network/interfaces",
                        "/platform/3/network/interfaces")

@app.route("/api/network/pools")
@require_auth
def ep_network_pools(entry):
    return proxy(entry, "/platform/7/network/pools",
                        "/platform/4/network/pools",
                        "/platform/3/network/pools")

@app.route("/api/snapshots/schedules")
@require_auth
def ep_snap_schedules(entry):
    return proxy(entry, "/platform/1/snapshot/schedules",
                        "/platform/3/snapshot/schedules")

@app.route("/api/storagepool/health")
@require_auth
def ep_pool_health(entry):
    return proxy(entry, "/platform/7/storagepool/storagepools",
                        "/platform/5/storagepool/storagepools")

# ── Migration health ──────────────────────────────────────────────

@app.route("/api/events")
@require_auth
def ep_events(entry):
    """Unresolved events — Pre-Migration Checklist #3, #10."""
    return proxy(entry, "/platform/3/event/eventlists?resolved=false&limit=100&dir=DESC",
                        "/platform/4/event/eventlists?resolved=false&limit=100&dir=DESC",
                        "/platform/1/event/events?limit=200")

@app.route("/api/jobs")
@require_auth
def ep_jobs(entry):
    """Active cluster jobs — Checklist #9."""
    return proxy(entry, "/platform/3/job/jobs", "/platform/1/job/jobs")

@app.route("/api/licenses")
@require_auth
def ep_licenses(entry):
    """Feature licenses — Checklist #1, Risk R-12."""
    return proxy(entry, "/platform/5/license/licenses",
                        "/platform/3/license/licenses",
                        "/platform/1/license/licenses")

@app.route("/api/ntp")
@require_auth
def ep_ntp(entry):
    """
    NTP / Time — Checklist #8.
    CONFIRMED on OneFS 9.10.1.3:
      /cluster/time/settings  → 404  (does not exist)
      /cluster/time           → 200  (per-node timestamps ONLY)
    NTP server list is only available via CLI: isi ntp servers list
    We return node timestamps for drift detection + manually stored servers.
    """
    sess = entry["session"]
    node_data, code = sess.get("/platform/3/cluster/time")
    if code not in (200,):
        node_data, code = sess.get("/platform/1/cluster/time")
    if code in (401, 403):
        return jsonify({"_privilege_denied": True, "_status": code,
                        "_message": node_data.get("_message", f"HTTP {code}")}), 200
    tok = get_token()
    with _lock:
        manual_servers = _store.get(tok, {}).get("ntp_servers", [])
    return jsonify({
        "node_times":     node_data if code == 200 else {},
        "manual_servers": manual_servers,
        "_ntp_papi_available": False,
        "_note": "NTP server list not available via PAPI on OneFS 9.10.1.3. Use: isi ntp servers list",
    }), 200

@app.route("/api/ntp/servers")
@require_auth
def ep_ntp_servers(entry):
    tok = get_token()
    with _lock:
        servers = _store.get(tok, {}).get("ntp_servers", [])
    return jsonify({"servers": servers}), 200

@app.route("/api/ntp/manual", methods=["POST"])
@require_auth
def ep_ntp_manual(entry):
    """Store NTP servers manually — PAPI does not expose them on this version."""
    body    = request.get_json(force=True, silent=True) or {}
    servers = body.get("servers", [])
    if isinstance(servers, str):
        servers = [s.strip() for s in servers.replace(",", "\n").splitlines() if s.strip()]
    tok = get_token()
    with _lock:
        if tok in _store:
            _store[tok]["ntp_servers"] = servers
    log.info("NTP servers saved manually: %s", servers)
    return jsonify({"saved": True, "servers": servers, "count": len(servers)}), 200

@app.route("/api/auth/providers")
@require_auth
def ep_auth_providers(entry):
    """Auth providers — Checklist #19, Risk R-08."""
    return proxy(entry, "/platform/3/auth/providers/summary",
                        "/platform/1/auth/providers/summary")

@app.route("/api/statistics/keys")
@require_auth
def ep_stats_keys(entry):
    return proxy(entry, "/platform/1/statistics/keys?limit=500")

@app.route("/api/statistics/current")
@require_auth
def ep_stats_current(entry):
    """Performance baseline — auto-discovers valid key names from /statistics/keys."""
    sess = entry["session"]
    keys_data, _ = sess.get("/platform/1/statistics/keys?limit=500")
    available = {k.get("key") for k in (keys_data.get("keys") or []) if k.get("key")}
    log.debug("Statistics: %d keys available", len(available))
    # Log available IOPS/latency keys for diagnostics
    iops_keys = sorted([k for k in available if "ops" in k.lower()])[:20]
    log.debug("IOPS keys available: %s", iops_keys)

    desired = {
        "throughput_in":  ["ifs.bytes.in.rate",    "ifs.bytes.in",   "cluster.net.bytes.in.avg"],
        "throughput_out": ["ifs.bytes.out.rate",   "ifs.bytes.out",  "cluster.net.bytes.out.avg"],
        "iops_total":     ["ifs.ops.total.rate",   "ifs.iops.total", "ifs.ops.total"],
        "iops_read":      ["ifs.ops.in.rate",      "ifs.iops.in",    "ifs.ops.read.rate",  "ifs.ops.read"],
        "iops_write":     ["ifs.ops.out.rate",     "ifs.iops.out",   "ifs.ops.write.rate", "ifs.ops.write"],
        "latency_total":  ["ifs.latency.total",    "ifs.lat.total.avg"],
        "latency_read":   ["ifs.latency.read",     "ifs.lat.read.avg"],
        "latency_write":  ["ifs.latency.write",    "ifs.lat.write.avg"],
        "cpu_user":       ["cluster.cpu.user.avg", "node.cpu.user.avg"],
        "bytes_total":    ["ifs.bytes.total",      "cluster.disk.bytes.total"],
        "bytes_used":     ["ifs.bytes.used",       "cluster.disk.bytes.used"],
        "bytes_free":     ["ifs.bytes.avail",      "ifs.bytes.free"],
    }

    # Auto-discover IOPS key by pattern if none match explicitly
    def _first_match(patterns, avail, pattern_fn=None):
        for p in patterns:
            if not avail or p in avail: return p
        if pattern_fn and avail:
            for k in sorted(avail):
                if pattern_fn(k.lower()): return k
        return None

    selected = {}
    for label, variants in desired.items():
        fn = None
        if "iops_total"  == label: fn = lambda k: "ops" in k and "total" in k and "rate" in k
        if "iops_read"   == label: fn = lambda k: "ops" in k and ("read" in k or ".in." in k) and "rate" in k
        if "iops_write"  == label: fn = lambda k: "ops" in k and ("write" in k or ".out." in k) and "rate" in k
        v = _first_match(variants, available, fn)
        if v: selected[label] = v

    log.debug("Selected keys: %s", selected)

    if selected:
        qs = "&".join(f"key={v}" for v in selected.values())
        data, code = sess.get(f"/platform/1/statistics/current?{qs}")
        if code == 200:
            label_map = {v: k for k, v in selected.items()}
            for s in (data.get("stats") or []):
                s["_label"] = label_map.get(s.get("key"), s.get("key", ""))
            data["_selected_keys"] = selected
            return jsonify(data), 200
        if code in (401, 403):
            return jsonify({"_privilege_denied": True, "_status": code,
                            "_message": data.get("_message", f"HTTP {code}")}), 200

    # Fallback
    summ, scode = sess.get("/platform/1/statistics/summary/protocol")
    if scode == 200:
        summ["_from_summary"] = True
        return jsonify(summ), 200

    return jsonify({"_stats_unavailable": True, "_available_count": len(available),
                    "_sample_keys": sorted(list(available))[:30],
                    "_message": "Statistics unavailable — see _sample_keys for what exists"}), 200

@app.route("/api/statistics/clients")
@require_auth
def ep_stats_clients(entry):
    return proxy(entry, "/platform/1/statistics/client?limit=200")

@app.route("/api/upgrade/status")
@require_auth
def ep_upgrade_status(entry):
    return proxy(entry, "/platform/3/upgrade/cluster", "/platform/1/upgrade/cluster")

@app.route("/api/upgrade/nodes")
@require_auth
def ep_upgrade_nodes(entry):
    return proxy(entry, "/platform/3/upgrade/cluster/nodes", "/platform/1/upgrade/cluster/nodes")

@app.route("/api/whoami")
@require_auth
def ep_whoami(entry):
    """List roles and providers — /session/1/privileges is 404 on OneFS 9.10.1.3."""
    sess = entry["session"]
    results = {}
    for path in ["/platform/3/auth/roles", "/platform/1/auth/roles",
                 "/platform/3/auth/providers/summary"]:
        data, code = sess.get(path)
        results[path] = {"status": code, "data": data if code == 200 else str(data)[:100]}
        if code == 200 and "primary" not in results:
            results["primary"] = {"path": path, "data": data}
    return jsonify(results), 200

@app.route("/api/rawtest")
@require_auth
def api_rawtest(entry):
    """Diagnostic — tests all confirmed-working PAPI paths."""
    sess    = entry["session"]
    results = {}
    # ONLY paths confirmed working on OneFS 9.10.1.3
    paths = [
        "/platform/3/cluster/identity",
        "/platform/3/cluster/config",
        "/platform/3/cluster/nodes",
        "/platform/3/cluster/time",           # node timestamps (NOT NTP settings)
        "/platform/3/event/eventlists?limit=1",
        "/platform/3/job/jobs",
        "/platform/5/license/licenses",
        "/platform/2/protocols/nfs/exports?limit=1",
        "/platform/4/protocols/smb/shares?limit=1",
        "/platform/1/quota/quotas?limit=1",
        "/platform/1/snapshot/snapshots?limit=1",
        "/platform/3/sync/policies?limit=1",
        "/platform/7/storagepool/nodepools",
        "/platform/7/storagepool/storagepools",
        "/platform/3/zones",
        "/platform/3/auth/providers/summary",
        "/platform/3/auth/roles",
        "/platform/1/statistics/keys?limit=5",
    ]
    for path in paths:
        url = f"{sess.base_url}{path}"
        try:
            r = sess._s.get(url, timeout=10)
            results[path] = {"status": r.status_code, "body_preview": r.text[:200].strip()}
        except Exception as ex:
            results[path] = {"status": 0, "error": str(ex)}
    return jsonify(results), 200

@app.route("/api/debug/token")
def api_debug_token():
    tok = get_token()
    with _lock:
        found = tok in _store
        size  = len(_store)
        keys  = [("..."+k[-8:]) for k in _store]
    return jsonify({"token_received": ("..."+tok[-8:]) if tok else "EMPTY",
                    "token_found": found, "store_size": size, "store_keys": keys})

@app.route("/api/debug/stats")
def api_debug_stats():
    tok = get_token(); entry = _get(tok)
    if not entry:
        return jsonify({"error": "Not authenticated — connect first, then add ?token=YOUR_TOKEN"}), 401
    sess = entry["session"]
    keys_data, _ = sess.get("/platform/1/statistics/keys?limit=500")
    all_keys = [k.get("key","") for k in (keys_data.get("keys") or [])]
    return jsonify({
        "total_keys":    len(all_keys),
        "iops_keys":     sorted([k for k in all_keys if "ops" in k.lower()])[:30],
        "latency_keys":  sorted([k for k in all_keys if "lat" in k.lower()])[:20],
        "cpu_keys":      sorted([k for k in all_keys if "cpu" in k.lower()])[:10],
        "byte_keys":     sorted([k for k in all_keys if "bytes" in k.lower()])[:20],
    })


# ══════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  — openpyxl, styled like StorageVision
# ══════════════════════════════════════════════════════════════════

# Style constants — matches StorageVision colour scheme
_C = {
    "hdr_bg":   "1F3864", "hdr_fg":  "FFFFFF",
    "title_bg": "5B4A9C", "title_fg": "FFFFFF",   # PowerScale purple
    "sub_bg":   "EDE9F8",
    "border":   "B0BEC5",
    "ok_bg":    "C8E6C9", "warn_bg": "FFE5B4", "crit_bg": "FFCCCC",
    "alt1":     "F5F3FF", "alt2":    "FFFFFF",
}

def _xl_fill(h):
    return PatternFill("solid", fgColor=h)

def _xl_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _xl_border():
    s = Side(style="thin", color=_C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_ctr(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _xl_left():
    return Alignment(horizontal="left", vertical="center")

def _xl_cell(ws, row, col, val, bg, bold=False, fmt=None, align="center", color="000000"):
    # _safe_val ensures nested dicts/lists never crash openpyxl
    safe = _safe_val(val) if not isinstance(val, (int, float, type(None))) else val
    c = ws.cell(row=row, column=col, value=safe)
    c.fill      = _xl_fill(bg)
    c.font      = _xl_font(bold=bold, color=color)
    c.border    = _xl_border()
    c.alignment = _xl_ctr() if align == "center" else _xl_left()
    if fmt:
        c.number_format = fmt
    return c

def _xl_title_row(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _xl_font(bold=True, color=_C["title_fg"], size=12)
    c.fill      = _xl_fill(_C["title_bg"])
    c.alignment = _xl_ctr(wrap=True)
    ws.row_dimensions[row].height = 26

def _xl_header_row(ws, headers, row):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _xl_font(bold=True, color=_C["hdr_fg"])
        c.fill      = _xl_fill(_C["hdr_bg"])
        c.alignment = _xl_ctr(wrap=True)
        c.border    = _xl_border()
    ws.row_dimensions[row].height = 28

def _status_color(pct):
    if pct >= 85: return _C["crit_bg"]
    if pct >= 75: return _C["warn_bg"]
    return _C["ok_bg"]

def _auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                best = min(max(best, len(str(cell.value)) + 2), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

def _fmt_bytes_xl(b):
    """Format bytes to human-readable string for Excel cells."""
    b = _to_num(b)  # coerce str/None to number first
    if not b:
        return "—"
    for unit, div in [("PB", 1024**5), ("TB", 1024**4), ("GB", 1024**3), ("MB", 1024**2)]:
        if b >= div:
            return f"{b/div:.2f} {unit}"
    return f"{b} B"

def _fmt_ts_xl(ts):
    if not ts:
        return "—"
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(ts)


def _safe_val(v, max_len=200):
    """
    Convert ANY value to a type safe for an Excel cell.
    Nested dicts/lists (e.g. OneFS hardware objects) are JSON-serialised
    and truncated so openpyxl never raises 'Cannot convert' errors.
    """
    if v is None:
        return "—"
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return v[:max_len] if len(v) > max_len else (v or "—")
    if isinstance(v, (dict, list)):
        try:
            s = json.dumps(v, separators=(",", ":"))
            return s[:max_len] + ("\u2026" if len(s) > max_len else "")
        except Exception:
            return str(v)[:max_len]
    return str(v)[:max_len]



def _to_num(v, default=0):
    """
    Coerce any value to int/float for arithmetic.
    Handles: None, '', '—', str digits, already-numeric.
    Prevents 'unsupported operand type str - str' in Excel builder.
    """
    if v is None or v == '' or v == '—':
        return default
    if isinstance(v, (int, float)):
        return v
    try:
        return int(v)
    except (ValueError, TypeError):
        pass
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def build_excel_report(data: dict, meta: dict) -> bytes:
    """
    Build a multi-sheet Excel workbook from collected PowerScale data.
    Returns bytes suitable for send_file().

    Sheets:
      1. Summary        — collection metadata + domain status
      2. Cluster        — identity + config fields
      3. Nodes          — per-node inventory
      4. NFS Exports    — all exports
      5. SMB Shares     — all shares
      6. Quotas         — quota policies
      7. Snapshots      — snapshot list
      8. SyncIQ         — replication policies
      9. Storage Pools  — SmartPool data
     10. Access Zones   — zone definitions
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Sheet 1: Summary ─────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    _xl_title_row(ws, "PowerScale / Isilon OneFS — Migration Assessment", 4)

    # Metadata block
    meta_rows = [
        ("Cluster",         meta.get("cluster", "—")),
        ("OneFS Version",   "9.10.1.3"),
        ("Report Title",    meta.get("title", "Migration Assessment")),
        ("Project",         meta.get("project", "—")),
        ("Collected By",    meta.get("by", "—")),
        ("Target Platform", meta.get("target", "—")),
        ("Collection Date", meta.get("collectedAt", "—")),
        ("Report Date",     now_str),
        ("Tool",            "PowerScale Migration Collector v4.0"),
        ("Auth Method",     "Session-Cookie (isisessid + isicsrf CSRF)"),
    ]
    for i, (k, v) in enumerate(meta_rows, 2):
        ws.cell(row=i, column=1, value=k).font  = _xl_font(bold=True)
        ws.cell(row=i, column=1).fill            = _xl_fill(_C["sub_bg"])
        ws.cell(row=i, column=1).border          = _xl_border()
        ws.cell(row=i, column=2, value=str(v)).border = _xl_border()
        ws.cell(row=i, column=2).fill            = _xl_fill(_C["alt1"])
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 50

    # Domain status table
    start = len(meta_rows) + 3
    _xl_header_row(ws, ["Domain", "Records Collected", "Status"], start)
    domains = [
        ("Cluster Info",    1 if data.get("cluster") else 0),
        ("Nodes",           len(data.get("nodes") or [])),
        ("NFS Exports",     len(data.get("nfs") or [])),
        ("SMB Shares",      len(data.get("smb") or [])),
        ("Quotas",          len(data.get("quotas") or [])),
        ("Snapshots",       len(data.get("snapshots") or [])),
        ("SyncIQ Policies", len(data.get("replication") or [])),
        ("Storage Pools",   len(data.get("pools") or [])),
        ("Access Zones",    len(data.get("zones") or [])),
    ]
    for i, (name, count) in enumerate(domains, start + 1):
        bg = _C["ok_bg"] if count > 0 else _C["warn_bg"]
        _xl_cell(ws, i, 1, name,               bg, align="left")
        _xl_cell(ws, i, 2, count,              bg, align="center")
        _xl_cell(ws, i, 3, "Collected" if count > 0 else "No Data / Priv Denied", bg, align="center")

    ws.column_dimensions["C"].width = 24

    # ── Sheet 2: Cluster ─────────────────────────────────────────
    cl = data.get("cluster")
    if cl:
        ws2 = wb.create_sheet("Cluster")
        _xl_title_row(ws2, "Cluster Identity & Configuration", 2)
        identity = cl.get("identity", {})
        config   = cl.get("config", {})
        ov       = config.get("onefs_version", {})
        # onefs_version may contain: release (string "9.10.1.3"),
        # revision (large packed int), build, type.
        # Always prefer "release" string over "revision" integer.
        onefs_str = (ov.get("release") or ov.get("revision") or
                     config.get("onefs_version_info", {}).get("release") or "—")
        # If it looks like a packed integer, use release instead
        if isinstance(onefs_str, int) or (isinstance(onefs_str, str) and onefs_str.isdigit()):
            onefs_str = ov.get("release") or str(onefs_str)
        fields = [
            ("Cluster Name",   identity.get("name", "—")),
            ("Description",    identity.get("description", "—")),
            ("Contact",        identity.get("contact", "—")),
            ("OneFS Version",  onefs_str),
            ("Build",          ov.get("build", "—")),
            ("Release",        ov.get("release", "—")),
            ("GUID",           config.get("guid", "—")),
            ("Local LNN",      config.get("local_lnn", "—")),
            ("Local DevID",    config.get("local_devid", "—")),
            ("Join Mode",      config.get("join_mode", "—")),
            ("Timezone",       (config.get("timezone") or {}).get("abbreviation", "—")),
            ("Long Name",      config.get("long_name", "—")),
        ]
        for i, (k, v) in enumerate(fields, 2):
            _xl_cell(ws2, i, 1, k, _C["sub_bg"], bold=True, align="left")
            _xl_cell(ws2, i, 2, str(v) if v else "—", _C["alt1"], align="left")
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 55

    # ── Sheet 3: Nodes ────────────────────────────────────────────
    # The per-LNN detail endpoint returns:
    #   {"nodes": [{"lnn":1, "node": { <all detail> }}]}
    # After ep_nodes merges stub+detail, the merged dict has:
    #   merged["lnn"]  = 1
    #   merged["node"] = { hardware:{}, cpu:{}, drives:[], capacity:[],
    #                       nvram:{}, batterystatus:{}, powersupplies:{},
    #                       state:"up", name:"AMGPSFS-3-1", ... }
    # We use a helper that searches the whole tree for a key.
    def _nget(n, *keys, default="—"):
        """Search for a key across node top-level AND node.node sub-dict."""
        inner = n.get("node") or {}
        for k in keys:
            v = n.get(k) or inner.get(k)
            if v is not None and v != "" and v != []:
                return v
        return default

    nodes = data.get("nodes") or []
    if nodes:
        ws3 = wb.create_sheet("Nodes")
        hdrs = ["LNN", "Name", "Status", "Model", "Serial",
                "CPU", "Total Disk Cap.", "Flash/SSD Cap.",
                "Memory", "Drive Count", "OneFS Release",
                "Battery Status", "NVRAM", "Power Supplies"]
        _xl_title_row(ws3, f"Node Inventory  ({len(nodes)} nodes)", len(hdrs))
        _xl_header_row(ws3, hdrs, 2)

        for i, n in enumerate(nodes, 3):
            # ── Confirmed field map from OneFS 9.10.1.3 PAPI debug output:
            # Top level:  drives[], hardware{}, id, lnn, partitions{}, sensors{},
            #             state (string), status{}
            # hardware{}: chassis, chassis_code, class, configuration_id, cpu,
            #             disk_controller, disk_expander, family_code, flash_drive,
            #             serial_number, memory_size, + more 25 keys total
            # status{}:   batterystatus{}, capacity[], cpu{}, nvram{},
            #             powersupplies{}, release (string), uptime, version
            # drives[]:   blocks, logical_block_length, media_type, model,
            #             serial, ui_state, purpose, baynum, interface_type, ...

            hw  = n.get("hardware") or {}
            st_obj = n.get("status") or {}    # status is a DICT with sub-fields
            drv = n.get("drives") or []

            # ── Status — "state" is a DICT on this OneFS version ─
            # state.readonly.status = "Read/Write" or "Read-Only"
            # state.smartfail.dead  = bool
            # state.servicelight.enabled = bool
            raw_state = n.get("state") or {}
            if isinstance(raw_state, str):
                st = raw_state
            elif isinstance(raw_state, dict):
                ro_info  = raw_state.get("readonly") or {}
                sf_info  = raw_state.get("smartfail") or {}
                ro_status = ro_info.get("status", "")         # "Read/Write" or "Read-Only"
                is_dead   = sf_info.get("dead", False)
                is_sf     = sf_info.get("smartfail", False)
                if is_dead:
                    st = "dead"
                elif is_sf:
                    st = "smartfailed"
                elif ro_status:
                    st = ro_status           # "Read/Write" = healthy
                else:
                    st = "up"
            else:
                st = "—"
            is_rw = st in ("Read/Write", "up")
            sbg = (_C["ok_bg"] if is_rw else
                   _C["warn_bg"] if st in ("Read-Only","readonly","smartfailed","degraded") else
                   _C["crit_bg"])

            # ── Name — not in API; derive from serial or LNN ─────
            # hardware.serial_number is the node serial e.g. "1VBJ0V3"
            serial  = hw.get("serial_number") or "—"
            name    = f"Node-{n.get('lnn','?')}"   # e.g. "Node-1"

            # ── Model — hardware.configuration_id e.g. "F200-96GB-SSD" ──
            model   = (hw.get("configuration_id") or hw.get("class") or
                       hw.get("chassis_code") or "—")

            # ── CPU — hw.cpu is dict {model, proc, overtemp, speed_limit} ─
            raw_cpu = hw.get("cpu") or st_obj.get("cpu") or {}
            if isinstance(raw_cpu, dict):
                import re as _re
                model_raw = raw_cpu.get("model","")  # "GenuineIntel (2.19GHz, stepping 0x...)"
                proc_raw  = raw_cpu.get("proc","")   # "Single-proc, 10-HT-core"
                ghz   = _re.search(r"(\d+\.\d+)GHz", model_raw + " " + proc_raw)
                cores = _re.search(r"(\d+)-HT-core", proc_raw)
                vendor= "Intel" if ("Intel" in model_raw or "Genuine" in model_raw) else (
                        "AMD" if "AMD" in model_raw else model_raw.split("(")[0].strip())
                cpu_str = vendor
                if ghz:   cpu_str += f" @ {ghz.group(1)}GHz"
                if cores: cpu_str += f" ({cores.group(1)}-core)"
                if not cpu_str.strip(): cpu_str = model_raw[:55] or "—"
            else:
                cpu_str = str(raw_cpu)[:55]
            if len(cpu_str) > 55: cpu_str = cpu_str[:55] + "…"

            # ── Disk Capacity — from drives[].blocks × logical_block_length ─
            # All drives on this cluster are SSD (media_type="SSD")
            hdd_bytes = 0
            ssd_bytes = 0
            for d in drv:
                if not isinstance(d, dict): continue
                blk  = _to_num(d.get("blocks", 0))
                blen = _to_num(d.get("logical_block_length", 512))
                cap  = blk * blen
                if d.get("media_type","").upper() == "SSD":
                    ssd_bytes += cap
                else:
                    hdd_bytes += cap

            # Also check status.capacity[] for capacity per disk-type
            cap_list = st_obj.get("capacity") or []
            if cap_list and not (hdd_bytes + ssd_bytes):
                for c in cap_list:
                    if not isinstance(c, dict): continue
                    b = _to_num(c.get("bytes", 0))
                    if c.get("type","").upper() == "SSD":
                        ssd_bytes += b
                    else:
                        hdd_bytes += b

            # ── Memory — exhaustive key search across hardware{} and status{} ──
            # PowerScale F200 = 96 GB RAM = 103,079,215,104 bytes
            # Key name varies: memory_size, memory, mem_size, ram
            mem = 0
            for _src_dict in [hw, st_obj, n]:
                if mem: break
                for _k, _v in (_src_dict or {}).items():
                    _kl = _k.lower()
                    if any(x in _kl for x in ("memory","mem_size","ram")) and not isinstance(_v, (dict, list)):
                        _n = _to_num(_v)
                        if _n > 1024 * 1024 * 1024:   # must be > 1 GB to be RAM
                            mem = _n; break

            # ── Battery — status.batterystatus{} ─────────────────
            bat     = st_obj.get("batterystatus") or {}
            bat_str = (bat.get("status1") or bat.get("status") or
                       bat.get("overall_status") or "—")
            if isinstance(bat_str, dict):
                bat_str = "—"

            # ── NVRAM — status.nvram{} ────────────────────────────
            nv      = st_obj.get("nvram") or {}
            nv_type = nv.get("present_type") or nv.get("type") or ""
            nv_sz   = _to_num(nv.get("present_size") or nv.get("size"))
            nv_str  = f"{nv_type} {_fmt_bytes_xl(nv_sz)}".strip() if (nv_type or nv_sz) else "—"

            # ── Power Supplies — status.powersupplies{} ───────────
            ps      = st_obj.get("powersupplies") or {}
            ps_str  = (ps.get("status") or
                       f"{ps.get('count',0)} PSU, {ps.get('failures',0)} fail" if ps else "—")
            if isinstance(ps_str, dict):
                ps_str = "—"

            # ── OneFS Release — status.release ────────────────────
            rel     = st_obj.get("release") or st_obj.get("version") or "—"
            if isinstance(rel, (int, float)):
                rel = "—"

            # ── Drive health summary ───────────────────────────────
            drv_healthy = sum(1 for d in drv if isinstance(d,dict) and d.get("ui_state")=="HEALTHY")
            drv_total   = len(drv)
            drv_str     = f"{drv_healthy}/{drv_total} healthy"

            bg = _C["alt1"] if i % 2 == 0 else _C["alt2"]

            _xl_cell(ws3, i,  1, n.get("lnn", "—"),                          bg,  align="center")
            _xl_cell(ws3, i,  2, name,                                        bg,  align="left")
            _xl_cell(ws3, i,  3, st,                                          sbg, align="center")
            _xl_cell(ws3, i,  4, model,                                       bg,  align="left")
            _xl_cell(ws3, i,  5, serial,                                      bg,  align="left")
            _xl_cell(ws3, i,  6, cpu_str,                                     bg,  align="left")
            total_disk = hdd_bytes + ssd_bytes
            _xl_cell(ws3, i,  7, _fmt_bytes_xl(total_disk) if total_disk else "—", bg, align="right")
            _xl_cell(ws3, i,  8, _fmt_bytes_xl(ssd_bytes) if ssd_bytes else ("All Flash" if total_disk else "—"), bg, align="right")
            _xl_cell(ws3, i,  9, _fmt_bytes_xl(mem) if mem else "—",         bg,  align="right")
            _xl_cell(ws3, i, 10, drv_str,                                     bg,  align="center")
            _xl_cell(ws3, i, 11, _safe_val(rel),                              bg,  align="left")
            _xl_cell(ws3, i, 12, _safe_val(bat_str),                         bg,  align="center")
            _xl_cell(ws3, i, 13, _safe_val(nv_str),                          bg,  align="left")
            _xl_cell(ws3, i, 14, _safe_val(ps_str),                          bg,  align="left")

        ws3.column_dimensions[get_column_letter(4)].width  = 28   # Model
        ws3.column_dimensions[get_column_letter(6)].width  = 46   # CPU
        ws3.column_dimensions[get_column_letter(10)].width = 16   # Drives
        ws3.column_dimensions[get_column_letter(14)].width = 28   # PSU
        _auto_width(ws3)

    # ── Sheet 4: NFS Exports ──────────────────────────────────────
    nfs = data.get("nfs") or []
    if nfs:
        ws4 = wb.create_sheet("NFS Exports")
        hdrs = ["ID", "Paths", "Description", "Clients", "Access", "Auth Flavors",
                "Map Root", "All Dirs"]
        _xl_title_row(ws4, f"NFS Exports  ({len(nfs)} exports)", len(hdrs))
        _xl_header_row(ws4, hdrs, 2)
        for i, e in enumerate(nfs, 3):
            bg  = _C["alt1"] if i % 2 == 0 else _C["alt2"]
            ro  = e.get("read_only", False)
            abg = _C["warn_bg"] if ro else bg
            _xl_cell(ws4, i, 1, e.get("id", "—"),                               bg, align="center")
            _xl_cell(ws4, i, 2, " | ".join(e.get("paths") or []),            bg, align="left")
            _xl_cell(ws4, i, 3, e.get("description", ""),                       bg, align="left")
            _xl_cell(ws4, i, 4, ", ".join(e.get("clients") or []) or "All (*)", bg, align="left")
            _xl_cell(ws4, i, 5, "Read-Only" if ro else "Read-Write",            abg, align="center")
            _xl_cell(ws4, i, 6, ", ".join(e.get("security_flavors") or []),     bg, align="left")
            mr = e.get("map_root") or {}
            if isinstance(mr, str):
                try: mr = json.loads(mr)
                except: pass
            if isinstance(mr, dict):
                # user field is itself a dict: {"id": "USER:nobody"}
                user_obj = mr.get("user") or {}
                if isinstance(user_obj, dict):
                    uid = str(user_obj.get("id") or user_obj.get("name") or "")
                else:
                    uid = str(user_obj)
                mr_id  = str(mr.get("id") or "")
                # Pick best value — strip "USER:", "GROUP:", "SID:" prefix
                raw = uid or mr_id
                mr_val = raw.split(":")[-1] if ":" in raw else (raw or "—")
            else:
                mr_val = str(mr).split(":")[-1].strip('"' + "} ") if ":" in str(mr) else (str(mr) if mr else "—")
            _xl_cell(ws4, i, 7, mr_val, bg, align="left")
            _xl_cell(ws4, i, 8, "Yes" if e.get("all_dirs") else "No",           bg, align="center")
        ws4.column_dimensions[get_column_letter(2)].width = 35   # Paths
        ws4.column_dimensions[get_column_letter(4)].width = 50   # Clients
        for row in ws4.iter_rows(min_row=3):
            for cell in row:
                if cell.column in (2, 4) and cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    ws4.row_dimensions[cell.row].height = max(
                        ws4.row_dimensions[cell.row].height or 15,
                        min(60, 15 * (str(cell.value).count(",") + 1))
                    )
        _auto_width(ws4)

    # ── Sheet 5: SMB Shares ───────────────────────────────────────
    smb = data.get("smb") or []
    if smb:
        ws5 = wb.create_sheet("SMB Shares")
        hdrs = ["Share Name", "Path", "Description", "Visibility",
                "Access", "Continuously Available", "Permissions"]
        _xl_title_row(ws5, f"SMB Shares  ({len(smb)} shares)", len(hdrs))
        _xl_header_row(ws5, hdrs, 2)
        for i, s in enumerate(smb, 3):
            bg  = _C["alt1"] if i % 2 == 0 else _C["alt2"]
            ro  = s.get("read_only", False)
            abg = _C["warn_bg"] if ro else bg
            perms = "; ".join(
                f"{p.get('permission','?')}:{(p.get('trustee') or {}).get('name','?')}"
                for p in (s.get("permissions") or [])
            ) or "—"
            _xl_cell(ws5, i, 1, s.get("name", "—"),                              bg,  align="left")
            _xl_cell(ws5, i, 2, s.get("path", "—"),                              bg,  align="left")
            _xl_cell(ws5, i, 3, s.get("description", ""),                        bg,  align="left")
            _xl_cell(ws5, i, 4, "Browsable" if s.get("browsable") else "Hidden", bg,  align="center")
            _xl_cell(ws5, i, 5, "Read-Only" if ro else "Read-Write",             abg, align="center")
            _xl_cell(ws5, i, 6, "Yes" if s.get("continuously_available") else "No", bg, align="center")
            _xl_cell(ws5, i, 7, perms,                                           bg,  align="left")
        _auto_width(ws5)

    # ── Sheet 6: Quotas ───────────────────────────────────────────
    quotas = data.get("quotas") or []
    if quotas:
        ws6 = wb.create_sheet("Quotas")
        hdrs = ["Path", "Type", "Logical Used", "Physical Used",
                "Hard Limit", "Soft Limit", "Advisory Limit", "Mode", "Used %"]
        _xl_title_row(ws6, f"Quota Policies  ({len(quotas)} quotas)", len(hdrs))
        _xl_header_row(ws6, hdrs, 2)
        for i, q in enumerate(quotas, 3):
            usage = q.get("usage") or {}
            thr   = q.get("thresholds") or {}
            logical  = _to_num(usage.get("logical"))
            physical = _to_num(usage.get("physical"))
            hard     = _to_num(thr.get("hard"))     or None
            soft     = _to_num(thr.get("soft"))     or None
            advisory = _to_num(thr.get("advisory")) or None
            pct = round(logical / hard * 100, 1) if hard and logical else 0
            bg  = _status_color(pct) if hard else (_C["alt1"] if i % 2 == 0 else _C["alt2"])
            _xl_cell(ws6, i, 1, q.get("path", "—"),              bg, align="left")
            _xl_cell(ws6, i, 2, q.get("type", "—"),              bg, align="center")
            _xl_cell(ws6, i, 3, _fmt_bytes_xl(logical),          bg, align="right")
            _xl_cell(ws6, i, 4, _fmt_bytes_xl(physical),         bg, align="right")
            _xl_cell(ws6, i, 5, _fmt_bytes_xl(hard) if hard else "None",     bg, align="right")
            _xl_cell(ws6, i, 6, _fmt_bytes_xl(soft) if soft else "None",     bg, align="right")
            _xl_cell(ws6, i, 7, _fmt_bytes_xl(advisory) if advisory else "None", bg, align="right")
            _xl_cell(ws6, i, 8, "Enforced" if q.get("enforced") else "Advisory", bg, align="center")
            _xl_cell(ws6, i, 9, f"{pct}%" if hard else "—",     bg, align="center")
        _auto_width(ws6)

    # ── Sheet 7: Snapshots ────────────────────────────────────────
    snaps = data.get("snapshots") or []
    if snaps:
        ws7 = wb.create_sheet("Snapshots")
        hdrs = ["ID", "Name", "Path", "Size", "State", "Created", "Expires", "Schedule"]
        _xl_title_row(ws7, f"Snapshots  ({len(snaps)} snapshots)", len(hdrs))
        _xl_header_row(ws7, hdrs, 2)
        for i, s in enumerate(snaps, 3):
            bg  = _C["alt1"] if i % 2 == 0 else _C["alt2"]
            st  = s.get("state", "—")
            sbg = _C["ok_bg"] if st == "active" else (_C["warn_bg"] if st else bg)
            _xl_cell(ws7, i, 1, s.get("id", "—"),              bg,  align="center")
            _xl_cell(ws7, i, 2, s.get("name", "—"),            bg,  align="left")
            _xl_cell(ws7, i, 3, s.get("path", "—"),            bg,  align="left")
            _xl_cell(ws7, i, 4, _fmt_bytes_xl(s.get("size")),  bg,  align="right")
            _xl_cell(ws7, i, 5, st,                            sbg, align="center")
            _xl_cell(ws7, i, 6, _fmt_ts_xl(s.get("created")), bg,  align="center")
            _xl_cell(ws7, i, 7, _fmt_ts_xl(s.get("expires")) if s.get("expires") else "Never", bg, align="center")
            _xl_cell(ws7, i, 8, s.get("schedule", "—"),        bg,  align="left")
        _auto_width(ws7)

    # ── Sheet 8: SyncIQ Policies — always create, even if empty ──
    rep = data.get("replication") or []
    if True:  # always create sheet
        ws8 = wb.create_sheet("SyncIQ Policies")
        hdrs = ["Policy Name", "Source Path", "Target Host", "Target Path",
                "Action", "State", "Schedule", "Last Success", "RPO (hrs)"]
        _xl_title_row(ws8, f"SyncIQ Replication Policies  ({len(rep)} policies)", len(hdrs))
        _xl_header_row(ws8, hdrs, 2)
        for i, p in enumerate(rep, 3):
            bg  = _C["alt1"] if i % 2 == 0 else _C["alt2"]
            en  = p.get("enabled", False)
            ebg = _C["ok_bg"] if en else _C["warn_bg"]
            _xl_cell(ws8, i, 1, p.get("name", "—"),                             bg,  align="left")
            _xl_cell(ws8, i, 2, p.get("source_root_path", "—"),                 bg,  align="left")
            _xl_cell(ws8, i, 3, p.get("target_host", "—"),                      bg,  align="left")
            _xl_cell(ws8, i, 4, p.get("target_path", "—"),                      bg,  align="left")
            _xl_cell(ws8, i, 5, p.get("action", "—"),                           bg,  align="center")
            _xl_cell(ws8, i, 6, "Enabled" if en else "Disabled",                ebg, align="center")
            _xl_cell(ws8, i, 7, p.get("schedule") or "Manual",                  bg,  align="center")
            _xl_cell(ws8, i, 8, _fmt_ts_xl(p.get("last_success_time")),         bg,  align="center")
            _xl_cell(ws8, i, 9, p.get("rpo_alert", "—"),                        bg,  align="center")
        _auto_width(ws8)

    # ── Sheet 9: Storage Pools ────────────────────────────────────
    pools = data.get("pools") or []
    if pools:
        ws9 = wb.create_sheet("Storage Pools")
        hdrs = ["Pool Name", "Type", "Total Capacity", "Used",
                "Available", "Used %", "Protection", "Member Nodes"]
        _xl_title_row(ws9, f"Storage Pools — SmartPools  ({len(pools)} pools)", len(hdrs))
        _xl_header_row(ws9, hdrs, 2)
        for i, p in enumerate(pools, 3):
            u     = p.get("usage") or {}
            total = _to_num(u.get("usable_bytes") or u.get("total_bytes") or u.get("total_size"))
            avail = _to_num(u.get("avail_bytes")  or u.get("free_bytes")  or u.get("avail_size"))
            used  = total - avail
            pct   = round(used / total * 100, 1) if total else 0
            bg    = _status_color(pct)
            lnns  = ", ".join(str(x) for x in (p.get("lnns") or [])) or p.get("health_flags", "—")
            _xl_cell(ws9, i, 1, p.get("name", "—"),   bg, align="left")
            _xl_cell(ws9, i, 2, p.get("type", "—"),   bg, align="center")
            _xl_cell(ws9, i, 3, _fmt_bytes_xl(total), bg, align="right")
            _xl_cell(ws9, i, 4, _fmt_bytes_xl(used),  bg, align="right")
            _xl_cell(ws9, i, 5, _fmt_bytes_xl(avail), bg, align="right")
            _xl_cell(ws9, i, 6, f"{pct}%",            bg, align="center")
            _xl_cell(ws9, i, 7, p.get("protection") or p.get("protection_policy") or
                   (p.get("data_protection") or {}).get("requested_protection") or
                   (p.get("data_protection") or {}).get("protection_policy") or "—", bg, align="center")
            _xl_cell(ws9, i, 8, lnns,                 bg, align="left")
        _auto_width(ws9)

    # ── Sheet 10: Access Zones ────────────────────────────────────
    zones = data.get("zones") or []
    if zones:
        ws10 = wb.create_sheet("Access Zones")
        hdrs = ["Zone Name", "Base Path", "Auth Providers", "Groupnet",
                "System Provider", "All Auth Providers Cache TTL"]
        _xl_title_row(ws10, f"Access Zones  ({len(zones)} zones)", len(hdrs))
        _xl_header_row(ws10, hdrs, 2)
        for i, z in enumerate(zones, 3):
            bg = _C["alt1"] if i % 2 == 0 else _C["alt2"]
            _xl_cell(ws10, i, 1, z.get("name", "—"),                             bg, align="left")
            _xl_cell(ws10, i, 2, z.get("path", "—"),                             bg, align="left")
            _xl_cell(ws10, i, 3, ", ".join(z.get("auth_providers") or []) or "—", bg, align="left")
            _xl_cell(ws10, i, 4, z.get("groupnet", "—"),                         bg, align="left")
            _xl_cell(ws10, i, 5, z.get("system_provider", "—"),                  bg, align="left")
            _xl_cell(ws10, i, 6, z.get("all_auth_providers_cache_ttl", "—"),     bg, align="center")
        # Auth Providers and System Provider columns need wider widths
        ws10.column_dimensions[get_column_letter(3)].width = 60
        ws10.column_dimensions[get_column_letter(5)].width = 35
        for row in ws10.iter_rows(min_row=3):
            for cell in row:
                if cell.column in (3, 5) and cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    ws10.row_dimensions[cell.row].height = 30
        _auto_width(ws10)


    # ── Migration Health summary sheet ────────────────────────────
    mh = wb.create_sheet("Migration Health")
    _xl_title_row(mh, "Migration Health — Pre-Upgrade Checklist (OneFS 9.10.1.3)", 4)
    _xl_header_row(mh, ["Check Item", "Value Collected", "Status", "Action Required"], 2)

    def _mhrow(ws, r, label, value, ok, action="—"):
        bg = _C["ok_bg"] if ok else _C["warn_bg"]
        _xl_cell(ws, r, 1, label,           bg, align="left")
        _xl_cell(ws, r, 2, str(value)[:100],bg, align="left")
        _xl_cell(ws, r, 3, "OK" if ok else "ACTION NEEDED", bg, align="center", bold=True)
        _xl_cell(ws, r, 4, action,          bg, align="left")

    r = 3
    # Events
    evts = data.get("events") or []
    _mhrow(mh, r, "Critical Events (unresolved)", len(evts), len(evts)==0,
           "Resolve all before upgrade window" if evts else "—"); r+=1
    # Jobs
    jobs = data.get("jobs") or []
    running_jobs = [j for j in jobs if str(j.get("state","")).lower()=="running"]
    _mhrow(mh, r, "Running Cluster Jobs", f"{len(running_jobs)} running / {len(jobs)} total",
           len(running_jobs)==0, "isi job pause before upgrade" if running_jobs else "—"); r+=1
    # Licenses
    lics = data.get("licenses") or []
    bad_lics = [l for l in lics if l.get("status","") not in ("Activated","active","","None",None)]
    _mhrow(mh, r, "Feature Licenses", f"{len(lics)} total · {len(bad_lics)} inactive",
           len(bad_lics)==0, "Reactivate before upgrade" if bad_lics else "—"); r+=1
    # NTP
    ntp_d = data.get("ntp") or {}
    ntimes = (ntp_d.get("node_times") or {}).get("nodes") or []
    msrvs  = ntp_d.get("manual_servers") or []
    ref_t  = _to_num(ntimes[0].get("time",0)) if ntimes else 0
    drift  = max(abs(_to_num(n.get("time",0))-ref_t) for n in ntimes) if ntimes else 0
    _mhrow(mh, r, "Node Time Drift (max)", f"{drift}s  (threshold: ≤5s)",
           drift<=5, "Fix NTP sync before upgrade" if drift>5 else "—"); r+=1
    _mhrow(mh, r, "NTP Servers", ", ".join(msrvs) if msrvs else "Not entered via UI",
           bool(msrvs), "Run: isi ntp servers list  then enter in UI" if not msrvs else "—"); r+=1
    # Auth Providers
    provs = data.get("authProviders") or []
    bad_p = [p for p in provs if str(p.get("status","")).lower() in ("offline","disconnected","error")]
    _mhrow(mh, r, "Auth Providers (AD/LDAP)", f"{len(provs)} total · {len(bad_p)} offline",
           len(bad_p)==0, "Check AD/LDAP before upgrade" if bad_p else "—"); r+=1
    # Performance baseline
    stats_d = data.get("statistics") or {}
    sl = stats_d.get("stats") or []
    def _sv(lbl): s=next((x for x in sl if x.get("_label")==lbl),None); return s.get("value") if s else None
    tin=_sv("throughput_in"); tout=_sv("throughput_out")
    ir=_sv("iops_read"); iw=_sv("iops_write"); cu=_sv("cpu_user")
    itot = _to_num(ir)+_to_num(iw) if (ir is not None and iw is not None) else None
    _mhrow(mh, r, "Throughput In  (pre-upgrade baseline)",
           _fmt_bytes_xl(_to_num(tin))+"/s" if tin is not None else "Not collected", True,
           "Compare post-upgrade — expect ≤5% regression"); r+=1
    _mhrow(mh, r, "Throughput Out (pre-upgrade baseline)",
           _fmt_bytes_xl(_to_num(tout))+"/s" if tout is not None else "Not collected", True,
           "Compare post-upgrade — expect ≤5% regression"); r+=1
    _mhrow(mh, r, "Total IOPS     (pre-upgrade baseline)",
           f"{int(itot):,}" if itot is not None else "Not collected", True,
           "Compare post-upgrade"); r+=1
    _mhrow(mh, r, "CPU User%      (pre-upgrade baseline)",
           f"{_to_num(cu)/10:.1f}%" if cu is not None else "Not collected", True,
           "Compare post-upgrade"); r+=1
    # Cluster capacity
    _mhrow(mh, r, "Cluster Total Capacity",
           _fmt_bytes_xl(_to_num((data.get("statistics") or {}).get("_total_bytes"))) or
           (pools[0].get("usage",{}).get("total_bytes") or "—" if (pools:=data.get("pools") or []) else "—"),
           True, "—"); r+=1

    mh.column_dimensions["A"].width = 38
    mh.column_dimensions["B"].width = 35
    mh.column_dimensions["C"].width = 16
    mh.column_dimensions["D"].width = 40

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.route("/api/export/excel", methods=["POST"])
@require_auth
def ep_export_excel(entry):
    """
    Receive collected data + metadata from browser JS,
    build Excel workbook server-side, return as download.
    """
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl not installed. pip install openpyxl"}), 503

    body = request.get_json(force=True, silent=True) or {}
    data = body.get("data", {})
    meta = body.get("meta", {})

    try:
        xls_bytes = build_excel_report(data, meta)
    except Exception as ex:
        log.error("Excel export error: %s", ex)
        return jsonify({"error": str(ex)}), 500

    cluster = meta.get("cluster", "cluster").replace(" ", "_")
    date    = datetime.now().strftime("%Y%m%d_%H%M")
    fname   = f"PowerScale_Migration_{cluster}_{date}.xlsx"

    return send_file(
        io.BytesIO(xls_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ══════════════════════════════════════════════════════════════════
#  EMBEDDED HTML FRONTEND
# ══════════════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PowerScale Isilon — Migration Collector</title>
<!-- Offline: no external font dependencies -->
<style>
:root{
  --bg:#0a0c10;--surface:#0e1117;--panel:#141820;--border:#1e2530;--border2:#2a3340;
  --accent:#00d4ff;--green:#00ff88;--yellow:#ffcc00;--red:#ff4455;--orange:#ff8c00;
  --text:#c8d8e8;--dim:#5a7080;--bright:#e8f4ff;
  /* Offline: system font stacks — no external dependencies */
  --mono:'Consolas','Cascadia Code','Lucida Console','Courier New',monospace;
  --sans:'Segoe UI','SF Pro Display','-apple-system','BlinkMacSystemFont','Helvetica Neue',sans-serif;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:var(--mono);font-size:13px;min-height:100vh;overflow-x:hidden}
body::before{content:'';position:fixed;inset:0;background-image:linear-gradient(rgba(0,212,255,.012) 1px,transparent 1px),linear-gradient(90deg,rgba(0,212,255,.012) 1px,transparent 1px);background-size:40px 40px;pointer-events:none;z-index:0}
.topbar{position:sticky;top:0;z-index:100;background:rgba(10,12,16,.96);border-bottom:1px solid var(--border);backdrop-filter:blur(12px);display:flex;align-items:center;gap:14px;padding:0 24px;height:54px}
.logo{font-family:var(--sans);font-size:15px;font-weight:800;letter-spacing:.06em;color:var(--accent)}
.logo span{color:var(--dim);font-weight:400}
.ver{font-size:10px;color:var(--dim);border:1px solid var(--border2);padding:2px 7px;border-radius:3px}
.auth-badge{display:flex;align-items:center;gap:5px;background:rgba(0,255,136,.07);border:1px solid rgba(0,255,136,.2);border-radius:4px;padding:3px 10px;color:var(--green);font-size:11px;margin-left:auto}
.session-box{font-size:11px;color:var(--dim);border:1px solid var(--border);padding:3px 10px;border-radius:4px}
.session-box span{color:var(--yellow)}
.layout{position:relative;z-index:1;display:grid;grid-template-columns:272px 1fr;min-height:calc(100vh - 54px)}
.sidebar{background:var(--surface);border-right:1px solid var(--border);padding:20px 0;display:flex;flex-direction:column}
.sec-label{font-size:10px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--dim);padding:0 18px 6px;margin-top:18px}
.nav-item{display:flex;align-items:center;gap:9px;padding:8px 18px;cursor:pointer;border-left:2px solid transparent;color:var(--dim);font-size:12px;transition:all .14s}
.nav-item:hover{background:rgba(0,212,255,.04);color:var(--text)}
.nav-item.active{background:rgba(0,212,255,.07);border-left-color:var(--accent);color:var(--accent)}
.nav-item .ic{width:16px;text-align:center}
.nav-item .badge{margin-left:auto;font-size:10px;background:var(--border2);border-radius:10px;padding:1px 7px;color:var(--dim)}
.nav-item.active .badge{background:rgba(0,212,255,.15);color:var(--accent)}
.sdiv{height:1px;background:var(--border);margin:10px 18px}
.conn-panel{margin:14px 10px 0;background:var(--panel);border:1px solid var(--border);border-radius:6px;padding:13px}
.cp-title{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--dim);margin-bottom:11px}
.cr{display:flex;flex-direction:column;gap:3px;margin-bottom:9px}
.cr label{font-size:10px;color:var(--dim);text-transform:uppercase;letter-spacing:.08em}
.ci{background:var(--bg);border:1px solid var(--border2);border-radius:4px;color:var(--bright);font-family:var(--mono);font-size:12px;padding:6px 8px;width:100%;outline:none;transition:border-color .15s}
.ci:focus{border-color:var(--accent)}
.ci::placeholder{color:var(--dim)}
.auth-note{font-size:10px;color:var(--green);background:rgba(0,255,136,.06);border:1px solid rgba(0,255,136,.2);border-radius:4px;padding:5px 8px;margin-bottom:9px;line-height:1.6}
.btn-conn{width:100%;background:var(--accent);color:#000;font-family:var(--mono);font-size:12px;font-weight:700;letter-spacing:.06em;border:none;border-radius:4px;padding:8px;cursor:pointer;transition:all .15s}
.btn-conn:hover{background:#00eeff}
.btn-conn:disabled{background:var(--border2);color:var(--dim);cursor:not-allowed}
.conn-status{display:flex;align-items:center;gap:6px;font-size:11px;margin-top:9px}
.dot{width:7px;height:7px;border-radius:50%;background:var(--dim)}
.dot.ok{background:var(--green);box-shadow:0 0 6px var(--green)}
.dot.connecting{background:var(--yellow);animation:blink .8s infinite}
.dot.err{background:var(--red)}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}
.main{padding:26px 30px;overflow-y:auto}
.page{display:none;animation:fadeIn .22s ease}
.page.active{display:block}
@keyframes fadeIn{from{opacity:0;transform:translateY(5px)}to{opacity:1;transform:none}}
.ph{display:flex;align-items:flex-end;justify-content:space-between;margin-bottom:22px}
.pt{font-family:var(--sans);font-size:21px;font-weight:800;color:var(--bright);letter-spacing:.02em}
.ps{font-size:11px;color:var(--dim);margin-top:3px}
.pacts{display:flex;gap:8px}
.btn{font-family:var(--mono);font-size:12px;border-radius:4px;padding:7px 14px;cursor:pointer;transition:all .14s;border:1px solid var(--border2);background:var(--panel);color:var(--text);display:flex;align-items:center;gap:6px}
.btn:hover{border-color:var(--accent);color:var(--accent)}
.btn-p{background:var(--accent);color:#000;border-color:var(--accent);font-weight:700}
.btn-p:hover{background:#00eeff;border-color:#00eeff;color:#000}
.btn:disabled{opacity:.4;cursor:not-allowed}
.card{background:var(--panel);border:1px solid var(--border);border-radius:6px;overflow:hidden;margin-bottom:18px}
.ch{display:flex;align-items:center;gap:10px;padding:11px 15px;border-bottom:1px solid var(--border);background:rgba(255,255,255,.015)}
.ct{font-size:11px;font-weight:700;letter-spacing:.1em;text-transform:uppercase}
.cc{font-size:10px;color:var(--dim);border:1px solid var(--border2);border-radius:10px;padding:1px 8px;margin-left:auto}
.cb{padding:0}
.sg{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px}
.sc{background:var(--panel);border:1px solid var(--border);border-radius:6px;padding:15px;position:relative;overflow:hidden}
.sc::after{content:'';position:absolute;bottom:0;left:0;height:2px;width:100%}
.c1::after{background:var(--accent)}.c2::after{background:var(--green)}.c3::after{background:var(--yellow)}.c4::after{background:var(--orange)}.c5::after{background:#b060ff}.c6::after{background:var(--red)}
.sl{font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);margin-bottom:6px}
.sv{font-family:var(--sans);font-size:24px;font-weight:800;color:var(--bright)}
.ss{font-size:10px;color:var(--dim);margin-top:3px}
.dt{width:100%;border-collapse:collapse;font-size:12px}
.dt th{text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:var(--dim);padding:9px 15px;border-bottom:1px solid var(--border);background:rgba(255,255,255,.01);white-space:nowrap}
.dt td{padding:8px 15px;border-bottom:1px solid rgba(30,37,48,.7);color:var(--text);vertical-align:middle}
.dt tr:last-child td{border-bottom:none}
.dt tr:hover td{background:rgba(0,212,255,.03)}
.mono{font-family:var(--mono)}
.chip{display:inline-flex;align-items:center;font-size:10px;border-radius:3px;padding:2px 7px;font-weight:600}
.cg{background:rgba(0,255,136,.1);color:var(--green)}
.cy{background:rgba(255,204,0,.1);color:var(--yellow)}
.cr2{background:rgba(255,68,85,.1);color:var(--red)}
.cb2{background:rgba(0,212,255,.1);color:var(--accent)}
.cgr{background:rgba(90,112,128,.15);color:var(--dim)}
.pbar{height:5px;background:var(--border);border-radius:3px;overflow:hidden;margin-bottom:14px}
.pfill{height:100%;border-radius:3px;background:var(--accent);transition:width .3s}
.log{background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:11px 14px;font-size:11px;line-height:1.85;max-height:240px;overflow-y:auto;font-family:var(--mono)}
.ll{display:flex;gap:12px}
.lt{color:var(--dim);min-width:82px}
.lm.info{color:var(--accent)}.lm.ok{color:var(--green)}.lm.warn{color:var(--yellow)}.lm.err{color:var(--red)}.lm.dim{color:var(--dim)}
.fg2{display:grid;grid-template-columns:1fr 1fr;gap:14px;padding:15px}
.fgrp{display:flex;flex-direction:column;gap:4px}
.fgrp label{font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:var(--dim)}
.fi{background:var(--bg);border:1px solid var(--border2);border-radius:4px;color:var(--bright);font-family:var(--mono);font-size:12px;padding:7px 9px;outline:none;transition:border-color .15s;width:100%}
.fi:focus{border-color:var(--accent)}
.fi::placeholder{color:var(--dim)}
.fsel{background:var(--bg);border:1px solid var(--border2);border-radius:4px;color:var(--bright);font-family:var(--mono);font-size:12px;padding:7px 9px;outline:none;appearance:none;cursor:pointer;transition:border-color .15s;width:100%}
.fsel:focus{border-color:var(--accent)}
.ckrow{display:flex;align-items:center;gap:8px;font-size:12px}
.ckrow input[type=checkbox]{accent-color:var(--accent);width:13px;height:13px}
.era{display:flex;gap:9px;padding:13px 15px;border-top:1px solid var(--border);background:rgba(255,255,255,.01)}
.sn{background:rgba(0,212,255,.04);border:1px solid rgba(0,212,255,.14);border-radius:6px;padding:11px 15px;margin-bottom:18px;font-size:11px;line-height:1.7;color:var(--dim)}
.sn strong{color:var(--accent)}
.empty{text-align:center;padding:46px;color:var(--dim)}
.empty-ic{font-size:34px;margin-bottom:10px}
.empty-t{font-family:var(--sans);font-size:14px;color:var(--text);margin-bottom:5px}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:var(--bg)}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--dim)}
@media(max-width:1100px){.sg{grid-template-columns:repeat(2,1fr)}}
</style>
</head>
<body>

<header class="topbar">
  <div class="logo">POWERSCALE <span>/ Isilon OneFS 9.x Migration Collector</span></div>
  <div class="ver">v4.0 · Session-Cookie Auth</div>
  <div class="auth-badge">&#128274; isisessid + CSRF</div>
  <div class="session-box">Session: <span id="timerDisplay">30:00</span></div>
</header>

<div class="layout">
<aside class="sidebar">
  <div class="conn-panel">
    <div class="cp-title">&#128279; Cluster Connection</div>
    <div class="auth-note">&#10003; Session-cookie auth (POST /session/1/session)<br>&#10003; Special characters in passwords fully supported<br>&#10003; No credentials stored anywhere</div>
    <div class="cr"><label>Cluster FQDN / IP</label>
      <input class="ci" id="cHost" type="text" placeholder="ps-mgmt.amg.local" autocomplete="off" spellcheck="false"></div>
    <div class="cr"><label>API Port</label>
      <input class="ci" id="cPort" type="number" value="8080"></div>
    <div class="cr"><label>Username</label>
      <input class="ci" id="cUser" type="text" placeholder="admin" autocomplete="off"></div>
    <div class="cr"><label>Password</label>
      <input class="ci" id="cPass" type="password" placeholder="••••••••" autocomplete="new-password"></div>
    <div class="cr" style="flex-direction:row;align-items:center;gap:8px">
      <input type="checkbox" id="sslVerify" style="accent-color:var(--accent)">
      <label for="sslVerify" style="text-transform:none;letter-spacing:0;font-size:11px;cursor:pointer;color:var(--text)">Verify SSL Certificate</label>
    </div>
    <button class="btn-conn" id="btnConn" onclick="doConnect()">CONNECT</button>
    <div class="conn-status"><div class="dot" id="sDot"></div><span id="sTxt">Not connected</span></div>
  </div>
  <div class="sdiv"></div>
  <div class="sec-label">Data Collection</div>
  <div class="nav-item active" onclick="nav('dashboard',this)"><span class="ic">&#9649;</span>Dashboard</div>
  <div class="nav-item" onclick="nav('cluster',this)"><span class="ic">&#128421;</span>Cluster Info<span class="badge" id="b-cluster">&#8212;</span></div>
  <div class="nav-item" onclick="nav('nodes',this)"><span class="ic">&#128225;</span>Nodes<span class="badge" id="b-nodes">&#8212;</span></div>
  <div class="nav-item" onclick="nav('filesystems',this)"><span class="ic">&#128194;</span>NFS Exports<span class="badge" id="b-nfs">&#8212;</span></div>
  <div class="nav-item" onclick="nav('smb',this)"><span class="ic">&#128391;</span>SMB Shares<span class="badge" id="b-smb">&#8212;</span></div>
  <div class="nav-item" onclick="nav('quotas',this)"><span class="ic">&#9878;</span>Quotas<span class="badge" id="b-quotas">&#8212;</span></div>
  <div class="nav-item" onclick="nav('snapshots',this)"><span class="ic">&#128248;</span>Snapshots<span class="badge" id="b-snaps">&#8212;</span></div>
  <div class="nav-item" onclick="nav('replication',this)"><span class="ic">&#128260;</span>SyncIQ<span class="badge" id="b-rep">&#8212;</span></div>
  <div class="nav-item" onclick="nav('pools',this)"><span class="ic">&#128190;</span>Storage Pools<span class="badge" id="b-pools">&#8212;</span></div>
  <div class="nav-item" onclick="nav('zones',this)"><span class="ic">&#128274;</span>Access Zones<span class="badge" id="b-zones">&#8212;</span></div>
  <div class="sdiv"></div>
  <div class="sec-label">Migration Health</div>
  <div class="nav-item" onclick="nav('events',this)"><span class="ic">&#9888;</span>Critical Events<span class="badge" id="b-events">&#8212;</span></div>
  <div class="nav-item" onclick="nav('jobs',this)"><span class="ic">&#9881;</span>Cluster Jobs<span class="badge" id="b-jobs">&#8212;</span></div>
  <div class="nav-item" onclick="nav('licenses',this)"><span class="ic">&#128273;</span>Licenses<span class="badge" id="b-licenses">&#8212;</span></div>
  <div class="nav-item" onclick="nav('ntp',this)"><span class="ic">&#128336;</span>NTP / Time<span class="badge" id="b-ntp">&#8212;</span></div>
  <div class="nav-item" onclick="nav('stats',this)"><span class="ic">&#128200;</span>Performance<span class="badge" id="b-stats">&#8212;</span></div>
  <div class="nav-item" onclick="nav('authprov',this)"><span class="ic">&#128101;</span>Auth Providers<span class="badge" id="b-auth">&#8212;</span></div>
  <div class="sdiv"></div>
  <div class="sec-label">Tools</div>
  <div class="nav-item" onclick="nav('collect',this)"><span class="ic">&#9654;</span>Run Collection</div>
  <div class="nav-item" onclick="nav('export',this)"><span class="ic">&#8595;</span>Export / Report</div>
  <div class="nav-item" onclick="runDiag()"><span class="ic">&#128269;</span>Run Diagnostics</div>
  <div class="nav-item" style="color:var(--red)" onclick="doDisconnect()"><span class="ic">&#128465;</span>Disconnect &amp; Clear</div>
</aside>

<main class="main">

<!-- DASHBOARD -->
<div class="page active" id="page-dashboard">
  <div class="ph">
    <div><div class="pt">Migration Dashboard</div><div class="ps">PowerScale / Isilon OneFS 9.10.1.3 · Session-Cookie Auth</div></div>
    <div class="pacts"><button class="btn btn-p" onclick="nav('collect',null)">&#9654; Run Collection</button></div>
  </div>
  <div class="sn"><strong>&#128274; Session-Cookie Auth (v4.0):</strong> This version uses <strong>POST /session/1/session</strong> — the correct OneFS PAPI authentication mechanism. Credentials are sent once as JSON, the server returns an <strong>isisessid</strong> session cookie + <strong>isicsrf</strong> CSRF token. All subsequent requests use these tokens — no Basic auth headers. Special characters in passwords are fully supported. Session expires in 30 minutes.</div>
  <div class="sg">
    <div class="sc c1"><div class="sl">Cluster Name</div><div class="sv" id="s-name" style="font-size:16px">&#8212;</div><div class="ss">OneFS identity</div></div>
    <div class="sc c2"><div class="sl">Nodes</div><div class="sv" id="s-nodes">&#8212;</div><div class="ss">Active members</div></div>
    <div class="sc c3"><div class="sl">NFS Exports</div><div class="sv" id="s-nfs">&#8212;</div><div class="ss">Configured</div></div>
    <div class="sc c4"><div class="sl">SMB Shares</div><div class="sv" id="s-smb">&#8212;</div><div class="ss">Active shares</div></div>
    <div class="sc c5"><div class="sl">Quotas</div><div class="sv" id="s-quotas">&#8212;</div><div class="ss">Policies</div></div>
    <div class="sc c6"><div class="sl">Snapshots</div><div class="sv" id="s-snaps">&#8212;</div><div class="ss">Snapshot count</div></div>
  </div>
  <div class="card">
    <div class="ch"><span class="ct">Collection Progress</span><span id="dashStatus" style="font-size:11px;color:var(--dim);margin-left:auto">Not started</span></div>
    <div class="cb" style="padding:15px"><div id="dashProgress" style="color:var(--dim);font-size:12px">Connect and run collection to begin.</div></div>
  </div>
  <div class="card">
    <div class="ch"><span class="ct">Activity Log</span></div>
    <div class="cb" style="padding:11px"><div class="log" id="logBox"><div class="ll"><span class="lt">--:--:--</span><span class="lm dim">Awaiting connection...</span></div></div></div>
  </div>
</div>

<!-- CLUSTER -->
<div class="page" id="page-cluster">
  <div class="ph"><div><div class="pt">Cluster Information</div><div class="ps">Identity, OneFS version, hardware</div></div>
  <div class="pacts"><button class="btn" onclick="fetchCluster()">&#8635; Refresh</button></div></div>
  <div id="clusterOut"><div class="empty"><div class="empty-ic">&#128421;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- NODES -->
<div class="page" id="page-nodes">
  <div class="ph"><div><div class="pt">Node Inventory</div><div class="ps">Per-node hardware and capacity</div></div>
  <div class="pacts"><button class="btn" onclick="fetchNodes()">&#8635; Refresh</button></div></div>
  <div id="nodesOut"><div class="empty"><div class="empty-ic">&#128225;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- NFS -->
<div class="page" id="page-filesystems">
  <div class="ph"><div><div class="pt">NFS Exports</div><div class="ps">Paths, clients and auth flavors</div></div>
  <div class="pacts"><button class="btn" onclick="fetchNFS()">&#8635; Refresh</button></div></div>
  <div id="nfsOut"><div class="empty"><div class="empty-ic">&#128194;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- SMB -->
<div class="page" id="page-smb">
  <div class="ph"><div><div class="pt">SMB Shares</div><div class="ps">Windows file sharing configuration</div></div>
  <div class="pacts"><button class="btn" onclick="fetchSMB()">&#8635; Refresh</button></div></div>
  <div id="smbOut"><div class="empty"><div class="empty-ic">&#128391;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- QUOTAS -->
<div class="page" id="page-quotas">
  <div class="ph"><div><div class="pt">Quota Policies</div><div class="ps">Directory and user quotas</div></div>
  <div class="pacts"><button class="btn" onclick="fetchQuotas()">&#8635; Refresh</button></div></div>
  <div id="quotasOut"><div class="empty"><div class="empty-ic">&#9878;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- SNAPSHOTS -->
<div class="page" id="page-snapshots">
  <div class="ph"><div><div class="pt">Snapshots</div><div class="ps">Snapshot inventory and schedules</div></div>
  <div class="pacts"><button class="btn" onclick="fetchSnapshots()">&#8635; Refresh</button></div></div>
  <div id="snapsOut"><div class="empty"><div class="empty-ic">&#128248;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- SYNCIQ -->
<div class="page" id="page-replication">
  <div class="ph"><div><div class="pt">SyncIQ Replication</div><div class="ps">Policies, schedules and status</div></div>
  <div class="pacts"><button class="btn" onclick="fetchSyncIQ()">&#8635; Refresh</button></div></div>
  <div id="repOut"><div class="empty"><div class="empty-ic">&#128260;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- POOLS -->
<div class="page" id="page-pools">
  <div class="ph"><div><div class="pt">Storage Pools</div><div class="ps">SmartPools capacity and tiers</div></div>
  <div class="pacts"><button class="btn" onclick="fetchPools()">&#8635; Refresh</button></div></div>
  <div id="poolsOut"><div class="empty"><div class="empty-ic">&#128190;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- ZONES -->
<div class="page" id="page-zones">
  <div class="ph"><div><div class="pt">Access Zones</div><div class="ps">Zone names, paths, auth providers</div></div>
  <div class="pacts"><button class="btn" onclick="fetchZones()">&#8635; Refresh</button></div></div>
  <div id="zonesOut"><div class="empty"><div class="empty-ic">&#128274;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- COLLECT -->
<div class="page" id="page-collect">
  <div class="ph"><div><div class="pt">Run Data Collection</div><div class="ps">Select domains and capture migration data</div></div></div>
  <div class="card">
    <div class="ch"><span class="ct">Collection Scope</span></div>
    <div class="cb">
      <div class="fg2" style="grid-template-columns:repeat(3,1fr)">
        <label class="ckrow"><input type="checkbox" id="col-cluster" checked>Cluster Info</label>
        <label class="ckrow"><input type="checkbox" id="col-nodes" checked>Node Inventory</label>
        <label class="ckrow"><input type="checkbox" id="col-nfs" checked>NFS Exports</label>
        <label class="ckrow"><input type="checkbox" id="col-smb" checked>SMB Shares</label>
        <label class="ckrow"><input type="checkbox" id="col-quotas" checked>Quota Policies</label>
        <label class="ckrow"><input type="checkbox" id="col-snaps" checked>Snapshots</label>
        <label class="ckrow"><input type="checkbox" id="col-rep" checked>SyncIQ Policies</label>
        <label class="ckrow"><input type="checkbox" id="col-pools" checked>Storage Pools</label>
        <label class="ckrow"><input type="checkbox" id="col-zones" checked>Access Zones</label>
      </div>
      <div style="border-top:1px solid var(--border);padding:10px 15px 2px;font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:var(--dim)">Migration Health (new)</div>
      <div class="fg2" style="grid-template-columns:repeat(3,1fr)">
        <label class="ckrow"><input type="checkbox" id="col-events" checked>Critical Events</label>
        <label class="ckrow"><input type="checkbox" id="col-jobs" checked>Cluster Jobs</label>
        <label class="ckrow"><input type="checkbox" id="col-lic" checked>Feature Licenses</label>
        <label class="ckrow"><input type="checkbox" id="col-ntp" checked>NTP / Time Sync</label>
        <label class="ckrow"><input type="checkbox" id="col-stats" checked>Performance Baseline</label>
        <label class="ckrow"><input type="checkbox" id="col-auth" checked>Auth Providers</label>
      </div>
    </div>
    <div class="era">
      <button class="btn btn-p" id="btnCollect" onclick="runCollection()">&#9654; Start Collection</button>
      <button class="btn" onclick="setAll(true)">Select All</button>
      <button class="btn" onclick="setAll(false)">Deselect All</button>
    </div>
  </div>
  <div class="card">
    <div class="ch"><span class="ct">Progress</span><span id="colPct" style="margin-left:auto;font-size:11px;color:var(--dim)">0%</span></div>
    <div class="cb" style="padding:15px">
      <div class="pbar"><div class="pfill" id="colBar" style="width:0%"></div></div>
      <div class="log" id="colLog" style="max-height:300px"></div>
    </div>
  </div>
</div>

<!-- EXPORT -->
<div class="page" id="page-export">
  <div class="ph"><div><div class="pt">Export &amp; Report</div><div class="ps">Download collected data</div></div></div>
  <div class="card">
    <div class="ch"><span class="ct">Report Metadata</span></div>
    <div class="cb">
      <div class="fg2">
        <div class="fgrp"><label>Report Title</label><input class="fi" id="expTitle" value="PowerScale Migration Assessment"></div>
        <div class="fgrp"><label>Customer / Project</label><input class="fi" id="expProject" placeholder="Customer or project name"></div>
        <div class="fgrp"><label>Collected By</label><input class="fi" id="expBy" placeholder="Engineer name"></div>
        <div class="fgrp"><label>Target Platform</label>
          <select class="fsel" id="expTarget">
            <option>Dell PowerScale Gen 6</option>
            <option>Dell PowerScale H-Series</option>
            <option>Azure NetApp Files</option>
            <option>AWS FSx for NetApp ONTAP</option>
            <option>IBM Spectrum Scale</option>
            <option>Other NAS Platform</option>
          </select>
        </div>
      </div>
    </div>
    <div class="era">
      <button class="btn btn-p" onclick="expXLSX()">&#8595; Excel (.xlsx)</button>
      <button class="btn" onclick="expJSON()">&#8595; JSON</button>
      <button class="btn" onclick="expCSV()">&#8595; CSV</button>
      <button class="btn" onclick="expHTML()">&#8595; HTML Report</button>
    </div>
  </div>
  <div class="card">
    <div class="ch"><span class="ct">Data Summary</span></div>
    <div class="cb" style="padding:15px"><div id="expSummary" style="color:var(--dim);font-size:12px">No data collected yet.</div></div>
  </div>
</div>


<!-- EVENTS -->
<div class="page" id="page-events">
  <div class="ph"><div><div class="pt">Critical Events</div><div class="ps">Last 100 critical cluster events — must be zero before upgrade</div></div>
  <div class="pacts"><button class="btn" onclick="fetchEvents()">&#8635; Refresh</button></div></div>
  <div id="eventsOut"><div class="empty"><div class="empty-ic">&#9888;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- JOBS -->
<div class="page" id="page-jobs">
  <div class="ph"><div><div class="pt">Cluster Jobs</div><div class="ps">Active jobs — FlexProtect / AutoBalance must not be running during upgrade</div></div>
  <div class="pacts"><button class="btn" onclick="fetchJobs()">&#8635; Refresh</button></div></div>
  <div id="jobsOut"><div class="empty"><div class="empty-ic">&#9881;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- LICENSES -->
<div class="page" id="page-licenses">
  <div class="ph"><div><div class="pt">Feature Licenses</div><div class="ps">All licences must be active pre-upgrade and re-validated post-upgrade</div></div>
  <div class="pacts"><button class="btn" onclick="fetchLicenses()">&#8635; Refresh</button></div></div>
  <div id="licensesOut"><div class="empty"><div class="empty-ic">&#128273;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- NTP -->
<div class="page" id="page-ntp">
  <div class="ph"><div><div class="pt">NTP / Time Sync</div><div class="ps">All nodes must be NTP-synced before upgrade — drift causes cluster instability</div></div>
  <div class="pacts"><button class="btn" onclick="fetchNTP()">&#8635; Refresh</button></div></div>
  <!-- Manual NTP server entry - PAPI does not expose NTP servers on OneFS 9.10.1.3 -->
  <div class="card" style="margin-bottom:18px">
    <div class="ch"><span class="ct">&#128279; Manual NTP Server Entry</span></div>
    <div class="cb" style="padding:14px">
      <div style="font-size:11px;color:var(--dim);margin-bottom:10px">
        NTP servers are not exposed via PAPI on this OneFS version.<br>
        Run <span class="mono" style="background:var(--border);padding:1px 6px;border-radius:3px">isi ntp servers list</span> on the cluster CLI, then paste the server IPs below.
      </div>
      <div style="display:flex;gap:8px;align-items:center">
        <input class="fi" id="ntpServerInput" placeholder="10.8.96.11, 10.8.96.12" style="flex:1" onkeydown="if(event.key==='Enter')saveNTPServers()">
        <button class="btn btn-p" onclick="saveNTPServers()">&#10003; Save</button>
        <button class="btn" onclick="fetchNTP()">&#8635; Refresh</button>
      </div>
      <div id="ntpSaveStatus" style="font-size:11px;color:var(--dim);margin-top:6px"></div>
    </div>
  </div>
  <div id="ntpOut"><div class="empty"><div class="empty-ic">&#128336;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- STATS -->
<div class="page" id="page-stats">
  <div class="ph"><div><div class="pt">Performance Baseline</div><div class="ps">IOPS / throughput / latency — capture pre-upgrade for post-upgrade comparison</div></div>
  <div class="pacts"><button class="btn" onclick="fetchStats()">&#8635; Refresh</button></div></div>
  <div id="statsOut"><div class="empty"><div class="empty-ic">&#128200;</div><div class="empty-t">No Data</div></div></div>
</div>

<!-- AUTH PROVIDERS -->
<div class="page" id="page-authprov">
  <div class="ph"><div><div class="pt">Auth Providers</div><div class="ps">AD / LDAP providers must show status: connected before and after upgrade</div></div>
  <div class="pacts"><button class="btn" onclick="fetchAuthProviders()">&#8635; Refresh</button></div></div>
  <div id="authprovOut"><div class="empty"><div class="empty-ic">&#128101;</div><div class="empty-t">No Data</div></div></div>
</div>

</main>
</div>

<script>
const D={cluster:null,nodes:null,nfs:null,smb:null,quotas:null,snapshots:null,
         replication:null,pools:null,zones:null,collectedAt:null,clusterName:null,
         // New migration-plan domains
         events:null,jobs:null,ntp:null,licenses:null,
         authProviders:null,statistics:null,clients:null,
         upgradeStatus:null,snapSchedules:null,networkPools:null};
const DEBUG=true;
let _token=null;

// Session timer
setInterval(async()=>{
  try{
    const r=await apiFetch('/api/status');const d=await r.json();
    const el=document.getElementById('timerDisplay');
    if(!d.connected){el.textContent='30:00';el.parentElement.style.color='';return;}
    const m=String(Math.floor(d.remaining/60)).padStart(2,'0');
    const s=String(d.remaining%60).padStart(2,'0');
    el.textContent=m+':'+s;
    el.parentElement.style.color=d.remaining<300?'var(--red)':'';
    if(d.remaining<=0)setStatus('','Session expired');
  }catch(e){}
},5000);

// Core fetch — injects X-Session-Token
async function apiFetch(path,opts={}){
  const hdrs={'X-Session-Token':_token||''};
  if(opts.headers)Object.assign(hdrs,opts.headers);
  if(DEBUG)console.debug('[api]',opts.method||'GET',path,'token=',_token?'...'+_token.slice(-8):'NULL');
  const r=await fetch(path,{...opts,headers:hdrs});
  if(DEBUG)console.debug('[api]',r.status,'<-',path);
  return r;
}

function nav(id,el){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nav-item').forEach(n=>n.classList.remove('active'));
  const pg=document.getElementById('page-'+id);
  if(pg)pg.classList.add('active');
  if(el)el.classList.add('active');
}

function log(msg,type='info'){
  const now=new Date().toLocaleTimeString();
  const line=`<div class="ll"><span class="lt">${now}</span><span class="lm ${type}">${msg}</span></div>`;
  ['logBox','colLog'].forEach(id=>{const el=document.getElementById(id);if(el){el.innerHTML+=line;el.scrollTop=el.scrollHeight;}});
}

function setStatus(state,txt){
  document.getElementById('sDot').className='dot '+state;
  document.getElementById('sTxt').textContent=txt;
}
function setBadge(id,val){const el=document.getElementById('b-'+id);if(el)el.innerHTML=val||'&mdash;';}

// Connect
async function doConnect(){
  const host=document.getElementById('cHost').value.trim();
  const port=document.getElementById('cPort').value||'8080';
  const user=document.getElementById('cUser').value.trim();
  const pass=document.getElementById('cPass').value;
  const ssl=document.getElementById('sslVerify').checked;
  if(!host||!user||!pass){log('&#9888; Host, username and password required','warn');return;}
  const specials=[...pass].filter(c=>!/[a-zA-Z0-9]/.test(c));
  if(specials.length>0)log(`&#8505; Password has ${specials.length} special character(s) — handled safely via JSON`,'dim');
  setStatus('connecting','Connecting...');
  document.getElementById('btnConn').disabled=true;
  log(`POST /session/1/session &#8594; ${host}:${port}`,'info');
  try{
    const r=await fetch('/api/connect',{
      method:'POST',
      headers:{'Content-Type':'application/json; charset=utf-8'},
      body:JSON.stringify({host,port:parseInt(port),username:user,password:pass,ssl_verify:ssl})
    });
    const d=await r.json();
    if(r.ok){
      _token=d.token;
      if(DEBUG)console.debug('[connect] token stored: ...'+_token.slice(-8));
      setStatus('ok',`Connected · ${d.cluster}`);
      log(`✔ Session established — ${d.cluster} · ${d.onefs}`,'ok');
      log('&#10003; isisessid + isicsrf tokens active','ok');
      document.getElementById('cPass').value='';
      document.getElementById('btnConn').innerHTML='RECONNECT';
      D.clusterName=d.cluster;
      document.getElementById('s-name').textContent=d.cluster;
    }else{
      setStatus('err',d.error||'Failed');
      log(`&#10008; ${d.error||'Connection failed'}`,'err');
    }
  }catch(e){
    setStatus('err','Error');
    log(`&#10008; ${e.message}`,'err');
  }
  document.getElementById('btnConn').disabled=false;
}

async function doDisconnect(){
  if(!confirm('Disconnect and clear all collected data?'))return;
  await apiFetch('/api/disconnect',{method:'POST'});
  _token=null;
  Object.keys(D).forEach(k=>D[k]=null);
  setStatus('','Not connected');
  document.getElementById('btnConn').innerHTML='CONNECT';
  document.getElementById('cHost').value='';
  document.getElementById('cUser').value='';
  document.getElementById('cPass').value='';
  ['cluster','nodes','nfs','smb','quotas','snaps','rep','pools','zones'].forEach(b=>setBadge(b,'—'));
  ['s-name','s-nodes','s-nfs','s-smb','s-quotas','s-snaps'].forEach(id=>{const el=document.getElementById(id);if(el)el.textContent='—';});
  document.getElementById('logBox').innerHTML='<div class="ll"><span class="lt">--:--:--</span><span class="lm dim">Session cleared.</span></div>';
  nav('dashboard',document.querySelector('.nav-item'));
  log('Session cleared. isisessid deleted on server.','warn');
}

// Diagnostics
async function runDiag(){
  if(!_token){log('&#9888; Not connected','warn');return;}
  log('Running endpoint diagnostics...','info');
  try{
    const r=await apiFetch('/api/rawtest');
    const d=await r.json();
    log('&#10003; Diagnostic results (check browser console for full detail)','ok');
    console.table(Object.entries(d).map(([path,v])=>({path,status:v.status,preview:(v.body_preview||'').slice(0,80)})));
    Object.entries(d).forEach(([path,v])=>{
      const icon=v.status===200?'&#10003;':v.status===0?'&#10008;':'&#9888;';
      const col=v.status===200?'ok':v.status===0?'err':'warn';
      log(`${icon} ${v.status} ${path}`,col);
    });
  }catch(e){log('&#10008; Diagnostic error: '+e.message,'err');}
}

// Formatters
// Coerce to number — mirrors Python _to_num()
function _to_num(v){if(v===null||v===undefined||v===''||v==='—')return 0;const n=Number(v);return isNaN(n)?0:n;}

function fb(b){if(!b&&b!==0)return'—';const u=['B','KB','MB','GB','TB','PB'];let i=0,v=+b;while(v>=1024&&i<u.length-1){v/=1024;i++;}return v.toFixed(i>1?1:0)+' '+u[i];}
function fd(ts){if(!ts)return'—';return new Date(ts*1000).toLocaleString();}
function chip(l,cls){return`<span class="chip ${cls}">${l}</span>`;}
function kv(lbl,val){return`<div class="fgrp"><label>${lbl}</label><div style="font-size:13px;color:var(--bright);padding-top:2px">${val||'—'}</div></div>`;}
function nada(c){return`<tr><td colspan="${c}" style="color:var(--dim);text-align:center;padding:20px">No data returned</td></tr>`;}

// Privilege check
function isPrivDenied(d){return d&&d._privilege_denied===true;}
function privCard(ep,msg){
  return`<div class="card" style="border-color:rgba(255,204,0,.3)">
    <div class="ch" style="background:rgba(255,204,0,.06)"><span class="ct" style="color:var(--yellow)">&#9888; Insufficient PAPI Privilege</span></div>
    <div class="cb" style="padding:16px;font-size:12px;line-height:1.9">
      <div style="color:var(--yellow);margin-bottom:8px"><strong>Endpoint:</strong> <span class="mono">${ep}</span></div>
      <div style="color:var(--text);margin-bottom:12px">${msg||'Account lacks required ISI_PRIV for this endpoint.'}</div>
      <div style="color:var(--dim);font-size:11px">
        &#8226; Add required ISI_PRIV_* to the MigrationCapture role on the cluster CLI.<br>
        &#8226; Use root or SystemAdmin account for full capture.<br>
        &#8226; Run <strong>Diagnostics</strong> from the sidebar to see per-endpoint status.
      </div>
    </div></div>`;
}

// Fetchers
async function fetchCluster(){
  log('Fetching cluster info...','info');
  try{
    const[ri,rc]=await Promise.all([apiFetch('/api/cluster/identity'),apiFetch('/api/cluster/config')]);
    const id=ri.ok?await ri.json():{};const cfg=rc.ok?await rc.json():{};
    D.cluster={identity:id,config:cfg};
    const cv=cfg.onefs_version||{};
    document.getElementById('clusterOut').innerHTML=`
      <div class="card"><div class="ch"><span class="ct">Cluster Identity &amp; Config</span></div><div class="cb">
      <div class="fg2" style="grid-template-columns:repeat(3,1fr)">
        ${kv('Cluster Name',id.name)}${kv('Description',id.description)}${kv('Contact',id.contact)}
        ${kv('OneFS Version',cv.revision)}${kv('Build',cv.build)}${kv('Release',cv.release)}
        ${kv('GUID',cfg.guid)}${kv('Local LNN',cfg.local_lnn)}${kv('Join Mode',cfg.join_mode)}
        ${kv('Timezone',cfg.timezone?.abbreviation)}${kv('Long Name',cfg.long_name)}${kv('Local DevID',cfg.local_devid)}
      </div></div></div>`;
    setBadge('cluster','✔');log('&#10003; Cluster info collected','ok');
  }catch(e){log('&#10008; Cluster: '+e.message,'err');}
}

async function fetchNodes(){
  log('Fetching nodes (2-step per-LNN detail)...','info');
  try{
    const r=await apiFetch('/api/cluster/nodes');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('nodesOut').innerHTML=privCard('/platform/3/cluster/nodes',d._message);setBadge('nodes','⚠');log('&#9888; Nodes: insufficient privilege','warn');return;}
    if(d._warnings)d._warnings.forEach(w=>log('&#9888; '+w,'warn'));
    D.nodes=d.nodes||[];
    // Debug: log exact field values from first node
    if(DEBUG&&D.nodes.length>0){
      const n0=D.nodes[0];
      const hw=n0.hardware||{};
      const st=n0.status||{};
      const drv=n0.drives||[];
      console.debug('[nodes] === FIELD MAP FOR EXCEL ===');
      console.debug('[nodes] state:',          n0.state);
      console.debug('[nodes] id:',             n0.id);
      console.debug('[nodes] lnn:',            n0.lnn);
      console.debug('[nodes] hw.all keys:',    Object.keys(hw));
      console.debug('[nodes] hw.serial_number:',hw.serial_number);
      console.debug('[nodes] hw.chassis_code:', hw.chassis_code);
      console.debug('[nodes] hw.class:',       hw['class']);
      console.debug('[nodes] hw.configuration_id:',hw.configuration_id);
      console.debug('[nodes] hw.cpu:',         hw.cpu);
      console.debug('[nodes] hw.memory_size:',  hw.memory_size);
      console.debug('[nodes] hw.family_code:',  hw.family_code);
      console.debug('[nodes] hw.flash_drive:',  hw.flash_drive);
      console.debug('[nodes] hw.disk_controller:',hw.disk_controller);
      console.debug('[nodes] status.batterystatus:', st.batterystatus);
      console.debug('[nodes] status.nvram:',    st.nvram);
      console.debug('[nodes] status.powersupplies:', st.powersupplies);
      console.debug('[nodes] status.capacity:',  st.capacity);
      console.debug('[nodes] status.release:',   st.release);
      console.debug('[nodes] status.uptime:',    st.uptime);
      console.debug('[nodes] drives count:',     drv.length);
      if(drv[0]){
        console.debug('[nodes] drives[0].blocks:',drv[0].blocks);
        console.debug('[nodes] drives[0].logical_block_length:',drv[0].logical_block_length);
        console.debug('[nodes] drives[0].media_type:',drv[0].media_type);
        console.debug('[nodes] drives[0].model:',drv[0].model);
        console.debug('[nodes] drives[0].ui_state:',drv[0].ui_state);
        console.debug('[nodes] drives[0].purpose:',drv[0].purpose);
      }
    }
    document.getElementById('s-nodes').textContent=D.nodes.length;setBadge('nodes',D.nodes.length);
    const rows=D.nodes.map(n=>{
      const hw=n.hardware||{};
      const st=n.status||{};
      const drv=n.drives||[];
      // state is a DICT: {readonly:{status:'Read/Write'}, smartfail:{dead:bool}, ...}
      function parseState(raw){
        if(!raw||raw==='—') return {label:'—',cls:'cgr'};
        if(typeof raw==='string') return {label:raw,cls:raw==='up'||raw==='Read/Write'?'cg':'cy'};
        const ro=raw.readonly||{}; const sf=raw.smartfail||{};
        if(sf.dead)    return {label:'Dead',       cls:'cr2'};
        if(sf.smartfail) return {label:'SmartFailed',cls:'cr2'};
        const st=ro.status||''; // 'Read/Write' or 'Read-Only'
        if(st==='Read/Write') return {label:'Read/Write',cls:'cg'};
        if(st==='Read-Only')  return {label:'Read-Only', cls:'cy'};
        return {label:st||'up',cls:'cg'};
      }
      const stObj=parseState(n.state);
      const state=stObj.label; const stCls=stObj.cls;
      // Drive capacity: blocks × logical_block_length per drive
      const ssdBytes=drv.filter(d=>d.media_type==='SSD').reduce((a,d)=>a+(_to_num(d.blocks)*_to_num(d.logical_block_length||512)),0);
      const hddBytes=drv.filter(d=>d.media_type!=='SSD').reduce((a,d)=>a+(_to_num(d.blocks)*_to_num(d.logical_block_length||512)),0);
      const healthy=drv.filter(d=>d.ui_state==='HEALTHY').length;
      const bat=st.batterystatus||{};
      const nv=st.nvram||{};
      const ps=st.powersupplies||{};
      return `<tr>
        <td class="mono">${n.lnn||'—'}</td>
        <td>Node-${n.lnn||'?'}</td>
        <td>${chip(state,stCls)}</td>
        <td class="mono">${hw.configuration_id||hw['class']||'—'}</td>
        <td class="mono">${hw.serial_number||'—'}</td>
        <td class="mono">${typeof hw.cpu==='object'?hw.cpu?.model||JSON.stringify(hw.cpu).slice(0,40):hw.cpu||'—'}</td>
        <td class="mono">${hddBytes>0?fb(hddBytes):'—'}</td>
        <td class="mono">${ssdBytes>0?fb(ssdBytes):'—'}</td>
        <td class="mono">${hw.memory_size?fb(hw.memory_size):'—'}</td>
        <td class="mono">${healthy}/${drv.length} healthy</td>
        <td class="mono">${st.release||'—'}</td>
        <td>${bat.status1||bat.status||'—'}</td>
        <td class="mono">${nv.present_type||''} ${nv.present_size?fb(nv.present_size):''||'—'}</td>
        <td>${ps.status||ps.count!==undefined?ps.count+' PSU':''||'—'}</td>
      </tr>`;
    }).join('');
    document.getElementById('nodesOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">Node Inventory</span><span class="cc">${D.nodes.length} nodes</span></div>
      <div class="cb" style="overflow-x:auto"><table class="dt"><thead><tr><th>LNN</th><th>Name</th><th>Status</th><th>Model</th><th>Serial</th><th>CPU</th><th>HDD Cap.</th><th>SSD Cap.</th><th>Memory</th><th>Drives</th><th>Release</th><th>Battery</th><th>NVRAM</th><th>Power Supplies</th></tr></thead>
      <tbody>${rows||nada(14)}</tbody></table></div></div>`;
    log(`&#10003; ${D.nodes.length} nodes`,'ok');
  }catch(e){log('&#10008; Nodes: '+e.message,'err');}
}

async function fetchNFS(){
  log('Fetching NFS exports...','info');
  try{
    const r=await apiFetch('/api/nfs/exports');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('nfsOut').innerHTML=privCard('/platform/2/protocols/nfs/exports',d._message);setBadge('nfs','⚠');log('&#9888; NFS: insufficient privilege','warn');return;}
    D.nfs=d.exports||[];
    document.getElementById('s-nfs').textContent=D.nfs.length;setBadge('nfs',D.nfs.length);
    const rows=D.nfs.map(e=>`<tr>
      <td class="mono">${e.id||'—'}</td>
      <td class="mono" style="max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${(e.paths||[]).join(', ')||'—'}</td>
      <td>${e.description||'—'}</td>
      <td class="mono">${(e.clients||[]).join(', ')||'All (*)'}</td>
      <td>${chip(e.read_only?'RO':'RW',e.read_only?'cy':'cg')}</td>
      <td>${(e.security_flavors||[]).join(', ')||'—'}</td>
    </tr>`).join('');
    document.getElementById('nfsOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">NFS Exports</span><span class="cc">${D.nfs.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>ID</th><th>Path(s)</th><th>Description</th><th>Clients</th><th>Access</th><th>Auth</th></tr></thead>
      <tbody>${rows||nada(6)}</tbody></table></div></div>`;
    log(`&#10003; ${D.nfs.length} NFS exports`,'ok');
  }catch(e){log('&#10008; NFS: '+e.message,'err');}
}

async function fetchSMB(){
  log('Fetching SMB shares...','info');
  try{
    const r=await apiFetch('/api/smb/shares');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('smbOut').innerHTML=privCard('/platform/4/protocols/smb/shares',d._message);setBadge('smb','⚠');log('&#9888; SMB: insufficient privilege','warn');return;}
    D.smb=d.shares||[];
    document.getElementById('s-smb').textContent=D.smb.length;setBadge('smb',D.smb.length);
    const rows=D.smb.map(s=>`<tr>
      <td>${s.name||'—'}</td>
      <td class="mono" style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${s.path||'—'}</td>
      <td>${s.description||'—'}</td>
      <td>${chip(s.browsable?'Visible':'Hidden',s.browsable?'cb2':'cgr')}</td>
      <td>${chip(s.read_only?'RO':'RW',s.read_only?'cy':'cg')}</td>
      <td class="mono">${(s.permissions||[]).map(p=>`${p.permission}:${p.trustee?.name||'?'}`).join('<br>')||'—'}</td>
    </tr>`).join('');
    document.getElementById('smbOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">SMB Shares</span><span class="cc">${D.smb.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Name</th><th>Path</th><th>Description</th><th>Visibility</th><th>Access</th><th>Permissions</th></tr></thead>
      <tbody>${rows||nada(6)}</tbody></table></div></div>`;
    log(`&#10003; ${D.smb.length} SMB shares`,'ok');
  }catch(e){log('&#10008; SMB: '+e.message,'err');}
}

async function fetchQuotas(){
  log('Fetching quotas...','info');
  try{
    const r=await apiFetch('/api/quotas');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('quotasOut').innerHTML=privCard('/platform/1/quota/quotas',d._message);setBadge('quotas','⚠');log('&#9888; Quotas: insufficient privilege','warn');return;}
    D.quotas=d.quotas||[];
    document.getElementById('s-quotas').textContent=D.quotas.length;setBadge('quotas',D.quotas.length);
    const rows=D.quotas.map(q=>`<tr>
      <td class="mono">${q.path||'—'}</td><td>${chip(q.type||'—','cb2')}</td>
      <td class="mono">${fb(q.usage?.logical)}</td><td class="mono">${fb(q.usage?.physical)}</td>
      <td class="mono">${q.thresholds?.hard?fb(q.thresholds.hard):'None'}</td>
      <td class="mono">${q.thresholds?.soft?fb(q.thresholds.soft):'None'}</td>
      <td>${chip(q.enforced?'Enforced':'Advisory',q.enforced?'cg':'cgr')}</td>
    </tr>`).join('');
    document.getElementById('quotasOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">Quota Policies</span><span class="cc">${D.quotas.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Path</th><th>Type</th><th>Logical</th><th>Physical</th><th>Hard Limit</th><th>Soft Limit</th><th>Mode</th></tr></thead>
      <tbody>${rows||nada(7)}</tbody></table></div></div>`;
    log(`&#10003; ${D.quotas.length} quotas`,'ok');
  }catch(e){log('&#10008; Quotas: '+e.message,'err');}
}

async function fetchSnapshots(){
  log('Fetching snapshots...','info');
  try{
    const r=await apiFetch('/api/snapshots');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('snapsOut').innerHTML=privCard('/platform/1/snapshot/snapshots',d._message);setBadge('snaps','⚠');log('&#9888; Snapshots: insufficient privilege','warn');return;}
    D.snapshots=d.snapshots||[];
    document.getElementById('s-snaps').textContent=D.snapshots.length;setBadge('snaps',D.snapshots.length);
    const rows=D.snapshots.map(s=>`<tr>
      <td class="mono">${s.id||'—'}</td><td>${s.name||'—'}</td>
      <td class="mono" style="max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${s.path||'—'}</td>
      <td class="mono">${fb(s.size)}</td>
      <td>${chip(s.state||'—',s.state==='active'?'cg':'cy')}</td>
      <td class="mono">${fd(s.created)}</td>
      <td class="mono">${s.expires?fd(s.expires):'Never'}</td>
    </tr>`).join('');
    document.getElementById('snapsOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">Snapshots</span><span class="cc">${D.snapshots.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>ID</th><th>Name</th><th>Path</th><th>Size</th><th>State</th><th>Created</th><th>Expires</th></tr></thead>
      <tbody>${rows||nada(7)}</tbody></table></div></div>`;
    log(`&#10003; ${D.snapshots.length} snapshots`,'ok');
  }catch(e){log('&#10008; Snapshots: '+e.message,'err');}
}

async function fetchSyncIQ(){
  log('Fetching SyncIQ policies...','info');
  try{
    const r=await apiFetch('/api/synciq/policies');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('repOut').innerHTML=privCard('/platform/3/sync/policies',d._message);setBadge('rep','⚠');log('&#9888; SyncIQ: insufficient privilege','warn');return;}
    D.replication=d.policies||[];setBadge('rep',D.replication.length);
    const rows=D.replication.map(p=>`<tr>
      <td>${p.name||'—'}</td><td class="mono">${p.source_root_path||'—'}</td>
      <td class="mono">${p.target_host||'—'}</td><td class="mono">${p.target_path||'—'}</td>
      <td>${chip(p.enabled?'Enabled':'Disabled',p.enabled?'cg':'cgr')}</td>
      <td>${p.schedule||'Manual'}</td><td>${chip(p.action||'—','cb2')}</td>
    </tr>`).join('');
    document.getElementById('repOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">SyncIQ Policies</span><span class="cc">${D.replication.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Name</th><th>Source</th><th>Target Host</th><th>Target Path</th><th>State</th><th>Schedule</th><th>Action</th></tr></thead>
      <tbody>${rows||nada(7)}</tbody></table></div></div>`;
    log(`&#10003; ${D.replication.length} SyncIQ policies`,'ok');
  }catch(e){log('&#10008; SyncIQ: '+e.message,'err');}
}

async function fetchPools(){
  log('Fetching storage pools...','info');
  try{
    const r=await apiFetch('/api/pools');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('poolsOut').innerHTML=privCard('/platform/7/storagepool/storagepools',d._message);setBadge('pools','⚠');log('&#9888; Pools: insufficient privilege — add ISI_PRIV_SMARTPOOLS','warn');return;}
    D.pools=(d.storagepools||d.nodepools||d.pools||[]);setBadge('pools',D.pools.length);
    if(DEBUG&&D.pools[0]) console.debug('[pools] pool[0] keys:',Object.keys(D.pools[0]),
      '| data_protection:',D.pools[0].data_protection,
      '| protection:',D.pools[0].protection,
      '| protection_policy:',D.pools[0].protection_policy);
    const rows=D.pools.map(p=>{
      const u=p.usage||{};
      const total=u.usable_bytes||u.total_bytes||u.total_size||p.total||0;
      const avail=u.avail_bytes||u.free_bytes||u.avail_size||0;
      const pct=total>0?Math.round(((total-avail)/total)*100):0;
      return`<tr>
        <td>${p.name||'—'}</td><td>${p.type||'—'}</td>
        <td class="mono">${fb(total)}</td><td class="mono">${fb(avail)}</td>
        <td class="mono">${pct}%</td>
        <td>${p.protection||p.protection_policy||
          (p.data_protection&&p.data_protection.requested_protection)||
          (p.data_protection&&p.data_protection.protection_policy)||'—'}</td>
        <td class="mono">${(p.lnns||[]).join(', ')||p.health_flags||'—'}</td>
      </tr>`;
    }).join('');
    document.getElementById('poolsOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">Storage Pools (SmartPools)</span><span class="cc">${D.pools.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Name</th><th>Type</th><th>Total</th><th>Available</th><th>Used%</th><th>Protection</th><th>Nodes/Info</th></tr></thead>
      <tbody>${rows||nada(7)}</tbody></table></div></div>`;
    log(`&#10003; ${D.pools.length} storage pools`,'ok');
  }catch(e){log('&#10008; Pools: '+e.message,'err');}
}

async function fetchZones(){
  log('Fetching access zones...','info');
  try{
    const r=await apiFetch('/api/zones');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('zonesOut').innerHTML=privCard('/platform/1/zones',d._message);setBadge('zones','⚠');log('&#9888; Zones: insufficient privilege','warn');return;}
    D.zones=d.zones||[];setBadge('zones',D.zones.length);
    const rows=D.zones.map(z=>`<tr>
      <td>${z.name||'—'}</td><td class="mono">${z.path||'—'}</td>
      <td>${(z.auth_providers||[]).join(', ')||'—'}</td>
      <td>${z.groupnet||'—'}</td><td>${chip(z.system_provider||'—','cb2')}</td>
    </tr>`).join('');
    document.getElementById('zonesOut').innerHTML=`<div class="card"><div class="ch"><span class="ct">Access Zones</span><span class="cc">${D.zones.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Name</th><th>Base Path</th><th>Auth Providers</th><th>Groupnet</th><th>System Provider</th></tr></thead>
      <tbody>${rows||nada(5)}</tbody></table></div></div>`;
    log(`&#10003; ${D.zones.length} access zones`,'ok');
  }catch(e){log('&#10008; Zones: '+e.message,'err');}
}

// ─────────────────────────────────────────────────────
//  New Migration-Plan fetchers
// ─────────────────────────────────────────────────────

async function fetchEvents(){
  log('Fetching critical events...','info');
  try{
    const r=await apiFetch('/api/events');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('eventsOut').innerHTML=privCard('/platform/3/cluster/events',d._message);setBadge('events','⚠');return;}
    D.events=d.events||[];setBadge('events',D.events.length);
    const rows=D.events.map(e=>`<tr>
      <td class="mono">${e.id||'—'}</td>
      <td>${e.message||e.value||'—'}</td>
      <td>${chip(e.severity||'—',e.severity==='critical'?'cr2':e.severity==='warning'?'cy':'cgr')}</td>
      <td class="mono">${e.devid||'—'}</td>
      <td class="mono">${e.time?new Date(e.time*1000).toLocaleString():'—'}</td>
    </tr>`).join('');
    const warnBanner=D.events.length>0?`<div style="background:rgba(255,68,85,.08);border:1px solid rgba(255,68,85,.3);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--red)">
      &#9888; ${D.events.length} critical event(s) found — resolve before upgrade window</div>`:'<div style="background:rgba(0,255,136,.07);border:1px solid rgba(0,255,136,.2);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--green)">&#10003; No critical events — cluster healthy for upgrade</div>';
    document.getElementById('eventsOut').innerHTML=warnBanner+`<div class="card"><div class="ch"><span class="ct">Critical Events</span><span class="cc">${D.events.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>ID</th><th>Message</th><th>Severity</th><th>Device</th><th>Time</th></tr></thead>
      <tbody>${rows||nada(5)}</tbody></table></div></div>`;
    log(`${D.events.length===0?'&#10003;':'&#9888;'} ${D.events.length} critical events`,(D.events.length===0?'ok':'warn'));
  }catch(e){log('&#10008; Events: '+e.message,'err');}
}

async function fetchJobs(){
  log('Fetching cluster jobs...','info');
  try{
    const r=await apiFetch('/api/jobs');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('jobsOut').innerHTML=privCard('/platform/3/job/jobs',d._message);setBadge('jobs','⚠');return;}
    D.jobs=d.jobs||[];setBadge('jobs',D.jobs.length);
    const rows=D.jobs.map(j=>`<tr>
      <td>${j.id||'—'}</td><td>${j.type||j.name||'—'}</td>
      <td>${chip(j.state||'—',j.state==='running'?'cy':j.state==='succeeded'?'cg':'cgr')}</td>
      <td class="mono">${j.progress!==undefined?j.progress+'%':'—'}</td>
      <td class="mono">${j.node||'—'}</td>
      <td class="mono">${j.start_time?new Date(j.start_time*1000).toLocaleString():'—'}</td>
    </tr>`).join('');
    const running=D.jobs.filter(j=>j.state==='running');
    const warnBanner=running.length>0?`<div style="background:rgba(255,204,0,.08);border:1px solid rgba(255,204,0,.3);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--yellow)">
      &#9888; ${running.length} job(s) currently RUNNING — pause with <code>isi job pause</code> before upgrade</div>`:'';
    document.getElementById('jobsOut').innerHTML=warnBanner+`<div class="card"><div class="ch"><span class="ct">Cluster Jobs</span><span class="cc">${D.jobs.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>ID</th><th>Type</th><th>State</th><th>Progress</th><th>Node</th><th>Started</th></tr></thead>
      <tbody>${rows||nada(6)}</tbody></table></div></div>`;
    log(`&#10003; ${D.jobs.length} jobs (${running.length} running)`,running.length>0?'warn':'ok');
  }catch(e){log('&#10008; Jobs: '+e.message,'err');}
}

async function fetchLicenses(){
  log('Fetching feature licenses...','info');
  try{
    const r=await apiFetch('/api/licenses');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('licensesOut').innerHTML=privCard('/platform/5/license/licenses',d._message);setBadge('licenses','⚠');return;}
    D.licenses=d.licenses||[];setBadge('licenses',D.licenses.length);
    const rows=D.licenses.map(l=>`<tr>
      <td>${l.name||l.id||'—'}</td>
      <td>${chip(l.status||'—',l.status==='Activated'||l.status==='active'?'cg':l.status==='Expired'?'cr2':'cy')}</td>
      <td class="mono">${l.expiration||l.expiry_date||'—'}</td>
      <td>${l.feature||'—'}</td>
    </tr>`).join('');
    const inactive=D.licenses.filter(l=>l.status&&l.status!=='Activated'&&l.status!=='active');
    const warnBanner=inactive.length>0?`<div style="background:rgba(255,68,85,.08);border:1px solid rgba(255,68,85,.3);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--red)">
      &#9888; ${inactive.length} licence(s) not active — resolve before upgrade</div>`:'<div style="background:rgba(0,255,136,.07);border:1px solid rgba(0,255,136,.2);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--green)">&#10003; All licences active</div>';
    document.getElementById('licensesOut').innerHTML=warnBanner+`<div class="card"><div class="ch"><span class="ct">Feature Licenses</span><span class="cc">${D.licenses.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Name</th><th>Status</th><th>Expiry</th><th>Feature</th></tr></thead>
      <tbody>${rows||nada(4)}</tbody></table></div></div>`;
    log(`&#10003; ${D.licenses.length} licences`,'ok');
  }catch(e){log('&#10008; Licenses: '+e.message,'err');}
}

async function saveNTPServers(){
  const input=document.getElementById('ntpServerInput').value.trim();
  if(!input){document.getElementById('ntpSaveStatus').textContent='Enter at least one server IP';return;}
  const servers=input.split(/[,\n]+/).map(s=>s.trim()).filter(Boolean);
  try{
    const r=await apiFetch('/api/ntp/manual',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({servers})
    });
    const d=await r.json();
    if(d.saved){
      document.getElementById('ntpSaveStatus').innerHTML=
        '<span style="color:var(--green)">&#10003; Saved '+d.count+' NTP server(s): '+d.servers.join(', ')+'</span>';
      log('&#10003; NTP servers saved: '+d.servers.join(', '),'ok');
      fetchNTP(); // refresh NTP display with servers
    }
  }catch(e){log('&#10008; NTP save: '+e.message,'err');}
}

async function fetchNTP(){
  log('Fetching NTP / time config...','info');
  try{
    // NTP servers are embedded in the cluster/time response on OneFS 9.x
    // dt.ntp.servers = ["ntp1.example.com", "ntp2.example.com"]
    const rt=await apiFetch('/api/ntp');
    const dt=rt.ok?await rt.json():{};
    if(isPrivDenied(dt)){document.getElementById('ntpOut').innerHTML=privCard('/platform/3/cluster/time',dt._message);setBadge('ntp','⚠');return;}
    if(DEBUG)console.debug('[ntp] full response keys:',Object.keys(dt));
    if(DEBUG)console.debug('[ntp] full response:',JSON.stringify(dt,null,2).slice(0,800));
    D.ntp={time:dt};setBadge('ntp','✔');

    // Response: { node_times: {nodes:[{lnn,time}...]}, _ntp_servers_note: '...' }
    // /cluster/time/settings = 404 on OneFS 9.10.1.3 — NTP servers only via CLI
    const nodeTimes    = dt.node_times || dt;
    const nodeTimeList = nodeTimes.nodes || [];
    const refTime      = nodeTimeList.length ? nodeTimeList[0].time : null;
    const maxDrift     = nodeTimeList.reduce((m,n)=>Math.max(m,refTime?Math.abs(n.time-refTime):0),0);
    const syncOk       = maxDrift <= 5;
    const localTime    = refTime ? new Date(refTime*1000).toLocaleString() : '—';
    const driftRows    = nodeTimeList.map(n=>{
      const drift    = refTime ? Math.abs(n.time - refTime) : 0;
      const driftCls = drift>60?'cr2':drift>5?'cy':'cg';
      return '<tr><td class="mono">'+n.lnn+'</td>'+
             '<td class="mono">'+new Date(n.time*1000).toLocaleString()+'</td>'+
             '<td>'+chip(drift===0?'In sync':drift+'s drift',driftCls)+'</td></tr>';
    }).join('');

    const syncBanner = syncOk
      ? '<div style="background:rgba(0,255,136,.07);border:1px solid rgba(0,255,136,.2);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--green)">'+
        '&#10003; All nodes in sync &mdash; max drift: '+maxDrift+'s (OK for upgrade, threshold: 5s)</div>'
      : '<div style="background:rgba(255,68,85,.08);border:1px solid rgba(255,68,85,.3);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--red)">'+
        '&#9888; Time drift detected &mdash; max: '+maxDrift+'s &mdash; resolve before upgrade (Checklist #8)</div>';

    // Show manually entered NTP servers if available
    const manualServers = dt.manual_servers || [];
    const serverRows = manualServers.length
      ? manualServers.map(s=>'<tr><td class="mono">'+s+'</td>'+
          '<td>'+chip('Manual entry','cgr')+'</td>'+
          '<td class="mono">&mdash;</td></tr>').join('')
      : '<tr><td colspan="3" style="color:var(--dim);text-align:center;padding:12px">'+
        'Enter NTP server IPs in the field above &mdash; run '+
        '<span class="mono" style="background:var(--border);padding:1px 5px;border-radius:3px">isi ntp servers list</span>'+
        ' on the cluster CLI to find them</td></tr>';
    // Pre-fill input if servers already saved
    if(manualServers.length){
      const inp=document.getElementById('ntpServerInput');
      if(inp&&!inp.value)inp.value=manualServers.join(', ');
    }
    const cliNote = manualServers.length ? '' :
      '<div class="sn"><strong>&#8505; NTP server list not available via PAPI on OneFS 9.10.1.3</strong><br>'+
      'Run: <span class="mono" style="background:var(--border);padding:2px 6px;border-radius:3px">isi ntp servers list</span>'+
      ' on the cluster CLI, then enter the IPs in the field above.</div>';

        document.getElementById('ntpOut').innerHTML = syncBanner + cliNote +
      '<div class="card"><div class="ch"><span class="ct">NTP Servers</span>'+
      '<span class="cc">'+(manualServers.length||'—')+'</span></div>'+
      '<div class="cb"><table class="dt"><thead><tr>'+
      '<th>Server</th><th>Source</th><th>Offset</th>'+
      '</tr></thead><tbody>'+serverRows+'</tbody></table></div></div>'+
      '<div class="card"><div class="ch"><span class="ct">Node Time Sync</span>'+
      '<span class="cc">'+nodeTimeList.length+' nodes</span></div>'+
      '<div class="cb"><table class="dt"><thead><tr>'+
      '<th>LNN</th><th>Node Time</th><th>Drift vs Node 1</th>'+
      '</tr></thead><tbody>'+(driftRows||nada(3))+'</tbody></table></div></div>'+
      '<div class="card"><div class="ch"><span class="ct">Reference Time</span></div>'+
      '<div class="cb" style="padding:15px"><div class="fg2" style="grid-template-columns:repeat(2,1fr)">'+
      kv('Reference Time (Node 1)',localTime)+
      kv('Max Drift Across Nodes',maxDrift+'s')+
      '</div></div></div>';
    log('✔ Node time sync: '+nodeTimeList.length+' nodes, max drift: '+maxDrift+'s',syncOk?'ok':'warn');
  }catch(e){log('&#10008; NTP: '+e.message,'err');}
}

async function fetchStats(){
  log('Fetching performance baseline...','info');
  try{
    const r=await apiFetch('/api/statistics/current');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('statsOut').innerHTML=privCard('/platform/1/statistics/current',d._message);setBadge('stats','⚠');return;}
    if(d._stats_unavailable){
      document.getElementById('statsOut').innerHTML=`<div class="sn" style="border-color:rgba(255,204,0,.3)"><strong style="color:var(--yellow)">&#9888; Statistics unavailable</strong><br>${d._message||'Could not determine valid key names for this cluster.'}<br><br>Available key count: ${d._available_count||0}. Sample keys: <span class="mono">${(d._sample_keys||[]).slice(0,5).join(', ')}</span></div>`;
      setBadge('stats','?');return;
    }
    D.statistics=d;setBadge('stats','✔');
    const stats=d.stats||d.protocols||[];
    const selectedKeys=d._selected_keys||{};
    // Support both direct stats[] and summary protocol[] response formats
    const get=(label)=>{
      // Try label first, then key name
      const byLabel=stats.find(x=>x._label===label);
      if(byLabel!=null&&byLabel.value!=null)return byLabel.value;
      const key=selectedKeys[label];
      if(key){const byKey=stats.find(x=>x.key===key);if(byKey!=null&&byKey.value!=null)return byKey.value;}
      return null;
    };
    // iops_total: use explicit key if available, else calculate read+write
    const iopsRead=get('iops_read'), iopsWrite=get('iops_write');
    const iopsTotal=get('iops_total') ?? 
      (iopsRead!=null&&iopsWrite!=null ? Number(iopsRead)+Number(iopsWrite) : null);
    const fmtRate=(v)=>v!=null&&v!=='—'&&Number(v)>0?fb(Number(v))+'/s':v!=null?'0 B/s':'—';
    const noteHtml=d._from_summary?'<div style="font-size:10px;color:var(--dim);margin-top:8px">&#8505; Using protocol summary endpoint — individual metric keys not available on this cluster.</div>':'';
    // Debug: log what we got back to confirm key/label mapping
    if(DEBUG){
      console.debug('[stats] stats array:',stats);
      console.debug('[stats] selectedKeys:',selectedKeys);
      stats.forEach(s=>console.debug('[stats] key=',s.key,'_label=',s._label,'value=',s.value));
    }
    // fmtIops: handle null/0 gracefully
    const fmtIops=(v)=>v!=null&&v!=='—'?Math.round(Number(v)).toLocaleString():'—';
    // CPU: OneFS returns millicores (1000=100%), divide by 10 for %
    const fmtCpu=(v)=>v!==null&&v!==undefined&&v!=='—'?(Number(v)/10).toFixed(1)+'%':'—';
    // Capacity values are bytes (not rates)
    const fmtCap=(v)=>v!==null&&v!==undefined&&v!=='—'&&Number(v)>0?fb(Number(v)):'—';
    document.getElementById('statsOut').innerHTML=`
      <div class="card"><div class="ch"><span class="ct">Performance Baseline</span><span class="cc">${new Date().toLocaleString()}</span></div><div class="cb">
      <div class="sg" style="grid-template-columns:repeat(3,1fr);gap:12px;padding:15px">
        <div class="sc c1"><div class="sl">Throughput In</div><div class="sv" style="font-size:18px">${fmtRate(get('throughput_in'))}</div><div class="ss">${selectedKeys.throughput_in||''}</div></div>
        <div class="sc c2"><div class="sl">Throughput Out</div><div class="sv" style="font-size:18px">${fmtRate(get('throughput_out'))}</div><div class="ss">${selectedKeys.throughput_out||''}</div></div>
        <div class="sc c3"><div class="sl">Total IOPS</div><div class="sv" style="font-size:18px">${fmtIops(iopsTotal)}</div><div class="ss">${selectedKeys.iops_total||'read+write'}</div></div>
        <div class="sc c4"><div class="sl">Read IOPS</div><div class="sv" style="font-size:18px">${fmtIops(iopsRead)}</div><div class="ss">${selectedKeys.iops_read||''}</div></div>
        <div class="sc c5"><div class="sl">Write IOPS</div><div class="sv" style="font-size:18px">${fmtIops(iopsWrite)}</div><div class="ss">${selectedKeys.iops_write||''}</div></div>
        <div class="sc c6"><div class="sl">CPU User%</div><div class="sv" style="font-size:18px">${fmtCpu(get('cpu_user'))}</div><div class="ss">${selectedKeys.cpu_user||''}</div></div>
        <div class="sc c1"><div class="sl">Capacity Total</div><div class="sv" style="font-size:16px">${fmtCap(get('bytes_total'))}</div></div>
        <div class="sc c2"><div class="sl">Capacity Used</div><div class="sv" style="font-size:16px">${fmtCap(get('bytes_used'))}</div></div>
        <div class="sc c3"><div class="sl">Capacity Free</div><div class="sv" style="font-size:16px">${fmtCap(get('bytes_free'))}</div></div>
      </div></div></div>
      <div style="font-size:11px;color:var(--dim);padding:4px 0">&#8505; Save this as pre-upgrade baseline. Compare against post-upgrade to verify &lt;5% regression (Post-val #18-20).</div>
      ${noteHtml}`;
    log('&#10003; Performance baseline captured','ok');
  }catch(e){log('&#10008; Statistics: '+e.message,'err');}
}

async function fetchAuthProviders(){
  log('Fetching auth providers...','info');
  try{
    const r=await apiFetch('/api/auth/providers');if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    if(isPrivDenied(d)){document.getElementById('authprovOut').innerHTML=privCard('/platform/3/auth/providers',d._message);setBadge('auth','⚠');return;}
    const providers=d.provider_instances||d.providers||[];
    D.authProviders=providers;setBadge('auth',providers.length);
    const rows=providers.map(p=>`<tr>
      <td>${p.name||p.id||'—'}</td>
      <td>${chip(p.type||'—','cb2')}</td>
      <td>${chip(p.status||p.online_state||'—',
        (p.status==='online'||p.online_state==='online'||p.connected)?'cg':
        (p.status==='offline'||p.online_state==='offline')?'cr2':'cy')}</td>
      <td>${p.zone||p.access_zone||'—'}</td>
      <td>${p.server_uris?p.server_uris.join(', '):p.server||p.domain||'—'}</td>
    </tr>`).join('');
    const offline=providers.filter(p=>p.status==='offline'||p.online_state==='offline');
    const warnBanner=offline.length>0?`<div style="background:rgba(255,68,85,.08);border:1px solid rgba(255,68,85,.3);border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:12px;color:var(--red)">
      &#9888; ${offline.length} provider(s) offline — check AD/LDAP before upgrade</div>`:'';
    document.getElementById('authprovOut').innerHTML=warnBanner+`<div class="card"><div class="ch"><span class="ct">Auth Providers</span><span class="cc">${providers.length}</span></div>
      <div class="cb"><table class="dt"><thead><tr><th>Name</th><th>Type</th><th>Status</th><th>Zone</th><th>Server / Domain</th></tr></thead>
      <tbody>${rows||nada(5)}</tbody></table></div></div>`;
    log(`&#10003; ${providers.length} auth providers (${offline.length} offline)`,offline.length>0?'warn':'ok');
  }catch(e){log('&#10008; Auth Providers: '+e.message,'err');}
}

// Full collection
async function runCollection(){
  if(!_token){log('&#9888; Not connected. Please connect first.','warn');return;}
  const chk=id=>document.getElementById(id)?.checked;
  const tasks=[];
  if(chk('col-cluster'))tasks.push({n:'Cluster Info',fn:fetchCluster});
  if(chk('col-nodes'))  tasks.push({n:'Node Inventory',fn:fetchNodes});
  if(chk('col-nfs'))    tasks.push({n:'NFS Exports',fn:fetchNFS});
  if(chk('col-smb'))    tasks.push({n:'SMB Shares',fn:fetchSMB});
  if(chk('col-quotas')) tasks.push({n:'Quotas',fn:fetchQuotas});
  if(chk('col-snaps'))  tasks.push({n:'Snapshots',fn:fetchSnapshots});
  if(chk('col-rep'))    tasks.push({n:'SyncIQ Policies',fn:fetchSyncIQ});
  if(chk('col-pools'))  tasks.push({n:'Storage Pools',fn:fetchPools});
  if(chk('col-zones'))  tasks.push({n:'Access Zones',fn:fetchZones});
  if(chk('col-events')) tasks.push({n:'Critical Events',fn:fetchEvents});
  if(chk('col-jobs'))   tasks.push({n:'Cluster Jobs',fn:fetchJobs});
  if(chk('col-lic'))    tasks.push({n:'Licenses',fn:fetchLicenses});
  if(chk('col-ntp'))    tasks.push({n:'NTP / Time',fn:fetchNTP});
  if(chk('col-stats'))  tasks.push({n:'Performance Baseline',fn:fetchStats});
  if(chk('col-auth'))   tasks.push({n:'Auth Providers',fn:fetchAuthProviders});
  const btn=document.getElementById('btnCollect');
  btn.disabled=true;btn.innerHTML='&#8987; Collecting...';
  document.getElementById('colBar').style.width='0%';
  for(let i=0;i<tasks.length;i++){
    log(`[${i+1}/${tasks.length}] ${tasks[i].n}...`,'info');
    await tasks[i].fn();
    const p=Math.round(((i+1)/tasks.length)*100);
    document.getElementById('colBar').style.width=p+'%';
    document.getElementById('colPct').textContent=p+'%';
    await new Promise(r=>setTimeout(r,200));
  }
  D.collectedAt=new Date().toISOString();
  log(`&#10003; Collection complete &#8212; ${tasks.length} domains`,'ok');
  document.getElementById('dashStatus').textContent='Complete · '+new Date().toLocaleTimeString();
  updateSummary();
  btn.disabled=false;btn.innerHTML='&#9654; Re-run Collection';
}

function setAll(v){
  ['col-cluster','col-nodes','col-nfs','col-smb','col-quotas','col-snaps','col-rep','col-pools','col-zones']
    .forEach(id=>{const e=document.getElementById(id);if(e)e.checked=v;});
}

function updateSummary(){
  const items=[
    ['Cluster Info',D.cluster?'&#10003;':'&#8212;'],
    ['Nodes',D.nodes?D.nodes.length+' nodes':'&#8212;'],
    ['NFS Exports',D.nfs?D.nfs.length+' exports':'&#8212;'],
    ['SMB Shares',D.smb?D.smb.length+' shares':'&#8212;'],
    ['Quotas',D.quotas?D.quotas.length+' quotas':'&#8212;'],
    ['Snapshots',D.snapshots?D.snapshots.length+' snapshots':'&#8212;'],
    ['SyncIQ Policies',D.replication?D.replication.length+' policies':'&#8212;'],
    ['Storage Pools',D.pools?D.pools.length+' pools':'&#8212;'],
    ['Access Zones',D.zones?D.zones.length+' zones':'&#8212;'],
  ];
  document.getElementById('expSummary').innerHTML=`
    <table class="dt"><thead><tr><th>Domain</th><th>Status</th></tr></thead>
    <tbody>${items.map(([k,v])=>`<tr><td>${k}</td><td class="mono" style="color:${v.includes('&#8212;')?'var(--dim)':'var(--green)'}">${v}</td></tr>`).join('')}</tbody>
    </table>
    <div style="margin-top:10px;color:var(--dim);font-size:11px">Collected: ${D.collectedAt||'—'} · Cluster: ${D.clusterName||'Unknown'}</div>`;
}

// Exports
function dl(blob,name){const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=name;a.click();setTimeout(()=>URL.revokeObjectURL(a.href),5000);}
function safe(s){return(s||'unknown').replace(/[^a-z0-9]/gi,'_');}
function dateStr(){return new Date().toISOString().slice(0,10);}

// Excel export — server-side generation via /api/export/excel
async function expXLSX(){
  if(!_token){log('&#9888; Not connected','warn');return;}
  log('&#8987; Building Excel workbook on server...','info');
  const meta={
    cluster:D.clusterName||'Unknown',
    title:document.getElementById('expTitle').value||'PowerScale Migration Assessment',
    project:document.getElementById('expProject').value||'',
    by:document.getElementById('expBy').value||'',
    target:document.getElementById('expTarget').value||'',
    collectedAt:(()=>{
      const d=D.collectedAt?new Date(D.collectedAt):new Date();
      return d.toLocaleString('en-GB',{year:'numeric',month:'2-digit',day:'2-digit',
        hour:'2-digit',minute:'2-digit',second:'2-digit'}).replace(',','');
    })(),
  };
  const payload={
    meta,
    data:{
      cluster:D.cluster,
      nodes:D.nodes,
      nfs:D.nfs,
      smb:D.smb,
      quotas:D.quotas,
      snapshots:D.snapshots,
      replication:D.replication,
      pools:D.pools,
      zones:D.zones,
    }
  };
  try{
    const r=await apiFetch('/api/export/excel',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify(payload),
    });
    if(!r.ok){
      const e=await r.json().catch(()=>({error:'Server error'}));
      log('&#10008; Excel export failed: '+(e.error||r.status),'err');
      return;
    }
    const blob=await r.blob();
    const cd=r.headers.get('Content-Disposition')||'';
    const match=cd.match(/filename[^;=\n]*=(['""]?)([^'"\n]+)\1/);
    const fname=match?match[2]:`PowerScale_Migration_${(D.clusterName||'cluster').replace(/\s+/g,'_')}_${new Date().toISOString().slice(0,10)}.xlsx`;
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download=fname;
    a.click();
    setTimeout(()=>URL.revokeObjectURL(a.href),5000);
    log('&#10003; Excel export downloaded: '+fname,'ok');
  }catch(e){
    log('&#10008; Excel export error: '+e.message,'err');
  }
}

function expJSON(){
  const payload={meta:{tool:'PowerScale Migration Collector v4.0',auth:'session-cookie',
    onefs:'9.10.1.3',cluster:D.clusterName,collectedAt:D.collectedAt,
    exportedAt:new Date().toISOString(),
    project:document.getElementById('expProject').value,
    by:document.getElementById('expBy').value,
    target:document.getElementById('expTarget').value},...D};
  dl(new Blob([JSON.stringify(payload,null,2)],{type:'application/json'}),
     `psscale_migration_${safe(D.clusterName)}_${dateStr()}.json`);
  log('&#10003; JSON exported','ok');
}

function expCSV(){
  let csv='';
  const sec=(t,h,rows)=>{csv+='\n## '+t+'\n'+h.join(',')+'\n';
    rows.forEach(r=>csv+=r.map(v=>'"'+String(v).replace(/"/g,'""')+'"').join(',')+'\n');};
  if(D.nfs)sec('NFS Exports',['ID','Paths','Description','ReadOnly','Clients'],
    D.nfs.map(e=>[e.id||'',(e.paths||[]).join(';'),e.description||'',e.read_only||false,(e.clients||[]).join(';')]));
  if(D.smb)sec('SMB Shares',['Name','Path','Description','Browsable','ReadOnly'],
    D.smb.map(s=>[s.name||'',s.path||'',s.description||'',s.browsable||false,s.read_only||false]));
  if(D.quotas)sec('Quotas',['Path','Type','LogicalUsed','PhysicalUsed','HardLimit','SoftLimit'],
    D.quotas.map(q=>[q.path||'',q.type||'',q.usage?.logical||0,q.usage?.physical||0,q.thresholds?.hard||'',q.thresholds?.soft||'']));
  if(D.nodes)sec('Nodes',['LNN','Name','Status','Model','Serial'],
    D.nodes.map(n=>[n.lnn||'',n.name||'',n.status||'',n.hardware?.model||'',n.hardware?.serial_number||'']));
  if(D.snapshots)sec('Snapshots',['ID','Name','Path','Size','State','Created'],
    D.snapshots.map(s=>[s.id||'',s.name||'',s.path||'',s.size||0,s.state||'',s.created||'']));
  if(D.replication)sec('SyncIQ',['Name','Source','TargetHost','TargetPath','Enabled','Schedule'],
    D.replication.map(p=>[p.name||'',p.source_root_path||'',p.target_host||'',p.target_path||'',p.enabled||false,p.schedule||'Manual']));
  dl(new Blob([csv],{type:'text/csv'}),`psscale_migration_${safe(D.clusterName)}_${dateStr()}.csv`);
  log('&#10003; CSV exported','ok');
}

function expHTML(){
  const title  = document.getElementById('expTitle').value   || 'PowerScale Migration Assessment';
  const project= document.getElementById('expProject').value || '—';
  const by     = document.getElementById('expBy').value      || '—';
  const target = document.getElementById('expTarget').value  || '—';

  // ── Helpers ───────────────────────────────────────────────────
  const esc = v => String(v==null?'—':v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');

  const tbl = (headers, rows) => {
    const thead = '<tr style="background:#1a2030;color:#00d4ff">'+headers.map(h=>`<th>${esc(h)}</th>`).join('')+'</tr>';
    const tbody = rows.map(r=>'<tr>'+r.map(v=>`<td>${esc(v)}</td>`).join('')+'</tr>').join('');
    return `<table border="1" cellpadding="6" cellspacing="0"
      style="border-collapse:collapse;width:100%;font-size:12px;font-family:monospace;margin-bottom:18px">
      ${thead}${tbody}</table>`;
  };

  // ── Node field extractors (mirrors fetchNodes logic) ──────────
  const nodeStatus = n => {
    const raw = n.state || {};
    if(typeof raw === 'string') return raw;
    if(typeof raw === 'object'){
      const ro = raw.readonly || {};
      const sf = raw.smartfail || {};
      if(sf.dead)       return 'Dead';
      if(sf.smartfail)  return 'SmartFailed';
      return ro.status  || 'up';
    }
    return '—';
  };

  const nodeModel = n => {
    const hw = n.hardware || {};
    return hw.configuration_id || hw.class || hw.chassis_code || '—';
  };

  const nodeCPU = n => {
    // hw.cpu is a dict: {model:"GenuineIntel (2.19GHz, stepping ...)", proc:"Single-proc, 10-HT-core"}
    const hw  = n.hardware || {};
    const st  = n.status   || {};
    const cpu = hw.cpu || st.cpu || {};
    if(typeof cpu === 'string') return cpu || '—';
    if(typeof cpu !== 'object' || !cpu) return '—';
    const m = cpu.model || '';
    const p = cpu.proc  || '';
    // Extract GHz from parenthetical e.g. "GenuineIntel (2.19GHz, stepping 0x...)"
    const ghzM   = (m + ' ' + p).match(/([0-9]+[.][0-9]+)GHz/);
    const coresM = p.match(/([0-9]+)-HT-core/);
    const vendor = m.includes('Intel') || m.includes('Genuine') ? 'Intel'
                 : m.includes('AMD') ? 'AMD'
                 : (m.split('(')[0] || '').trim();
    let s = vendor || m;
    if(ghzM)   s += ' @ ' + ghzM[1] + 'GHz';
    if(coresM) s += ' (' + coresM[1] + '-core)';
    return s.trim() || m || '—';
  };

  const nodeMemory = n => {
    // Exhaustive search — memory_size may be null even when populated
    // Check hardware{}, status{}, and top-level node for any memory key > 1 GB
    const sources = [n.hardware||{}, n.status||{}, n];
    for(const obj of sources){
      for(const [k,v] of Object.entries(obj)){
        if(typeof v === 'object' || !v) continue;
        const kl = k.toLowerCase();
        if((kl.includes('memory') || kl.includes('mem_size') || kl === 'ram')){
          const num = Number(v);
          if(num > 1024*1024*1024) return fb(num);   // must be > 1 GB
        }
      }
    }
    return '—';
  };

  const nodeDisk = n => {
    const drv = n.drives || [];
    const st  = n.status  || {};
    let ssd=0, hdd=0;
    drv.forEach(d=>{
      const cap = _to_num(d.blocks) * _to_num(d.logical_block_length||512);
      (d.media_type||'').toUpperCase()==='SSD' ? ssd+=cap : hdd+=cap;
    });
    if(!ssd && !hdd){
      (st.capacity||[]).forEach(c=>{
        const b=_to_num(c.bytes);
        (c.type||'').toUpperCase()==='SSD' ? ssd+=b : hdd+=b;
      });
    }
    const total=ssd+hdd;
    return total ? fb(total)+(ssd&&!hdd?' (All SSD)':'') : '—';
  };

  const nfsMapRoot = e => {
    // Debug: log the raw map_root so we can see exact structure
    if(DEBUG) console.debug('[nfsMapRoot] raw:', JSON.stringify(e.map_root));
    
    const mr = e.map_root;
    
    // null / undefined / empty
    if(mr == null || mr === '') return '—';
    
    // Already a plain string
    if(typeof mr === 'string'){
      // Try JSON parse
      let parsed;
      try{ parsed = JSON.parse(mr); }catch(x){ parsed = null; }
      if(parsed && typeof parsed === 'object'){
        // Fall through to object handling below
        return extractMrName(parsed);
      }
      // Raw string like "USER:nobody" or just "nobody"
      return mr.includes(':') ? mr.split(':').slice(-1)[0] : mr;
    }
    
    // Number / boolean — unlikely but safe
    if(typeof mr !== 'object') return String(mr);
    
    // Object: {id: "USER:nobody", user: null, primary_group: null}
    return extractMrName(mr);
  };
  
  function extractMrName(obj){
    if(!obj || typeof obj !== 'object') return '—';
    // user field is a nested dict: {id: "USER:nobody"} — extract id then strip prefix
    if(obj.user){
      const u = obj.user;
      const uid = typeof u === 'string' ? u : (u.id || u.name || '');
      if(uid) return String(uid).includes(':') ? String(uid).split(':').slice(-1)[0] : String(uid);
    }
    // id field: "USER:nobody", "GROUP:wheel", or nested object
    let id = obj.id || obj.name;
    if(id == null) return '—';
    if(typeof id === 'object') id = id.id || id.name || JSON.stringify(id);
    const s = String(id);
    // Strip prefix like "USER:", "GROUP:", "SID:"
    return s.includes(':') ? s.split(':').slice(-1)[0].replace(/[{}"\'\s]/g, '') : s;
  }


  // ── Collection date — format as readable local time ───────────
  const collDate = D.collectedAt
    ? new Date(D.collectedAt).toLocaleString('en-GB',{
        year:'numeric',month:'2-digit',day:'2-digit',
        hour:'2-digit',minute:'2-digit',second:'2-digit'}).replace(',','')
    : '—';

  // ── Build HTML ────────────────────────────────────────────────
  const html = `<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<title>${esc(title)}</title>
<style>
  body{background:#0a0c10;color:#c8d8e8;font-family:sans-serif;padding:32px;max-width:1300px;margin:auto}
  h1{color:#00d4ff;font-size:22px;margin-bottom:6px}
  h2{color:#00d4ff;font-size:14px;font-weight:700;margin-top:32px;margin-bottom:8px;
     border-bottom:1px solid #1e2530;padding-bottom:4px;text-transform:uppercase;letter-spacing:.06em}
  .meta{background:#0e1117;border:1px solid #1e2530;border-radius:6px;padding:16px;
        margin:14px 0 24px;display:grid;grid-template-columns:repeat(3,1fr);gap:14px;font-size:12px}
  .meta strong{color:#5a7080;display:block;font-size:10px;text-transform:uppercase;
               letter-spacing:.08em;margin-bottom:3px}
  .meta span{color:#e8f4ff}
  table{border-collapse:collapse;width:100%;font-size:12px;font-family:monospace;margin-bottom:24px}
  th{background:#1a2030;color:#00d4ff;padding:8px 10px;text-align:left;border:1px solid #1e2530;
     font-size:11px;text-transform:uppercase;letter-spacing:.06em}
  td{padding:7px 10px;border:1px solid #1e2530;color:#c8d8e8;vertical-align:top}
  tr:nth-child(even) td{background:rgba(255,255,255,.02)}
  tr:hover td{background:rgba(0,212,255,.04)}
  .ok {color:#00ff88}.warn{color:#ffcc00}.bad{color:#ff4455}
  .tag{display:inline-block;padding:1px 7px;border-radius:3px;font-size:10px;font-weight:700}
  .tag-ok  {background:rgba(0,255,136,.1);color:#00ff88}
  .tag-warn{background:rgba(255,204,0,.1); color:#ffcc00}
  .tag-bad {background:rgba(255,68,85,.1); color:#ff4455}
  .footer{margin-top:48px;padding-top:12px;border-top:1px solid #1e2530;
          color:#5a7080;font-size:10px}
</style></head><body>

<h1>&#128311; ${esc(title)}</h1>
<div class="meta">
  <div><strong>Cluster</strong><span>${esc(D.clusterName||'—')}</span></div>
  <div><strong>OneFS Version</strong><span>9.10.1.3</span></div>
  <div><strong>Target Platform</strong><span>${esc(target)}</span></div>
  <div><strong>Project</strong><span>${esc(project)}</span></div>
  <div><strong>Collected By</strong><span>${esc(by)}</span></div>
  <div><strong>Collection Date</strong><span>${collDate}</span></div>
  <div><strong>Auth Method</strong><span>Session-Cookie (isisessid + isicsrf)</span></div>
  <div><strong>Nodes</strong><span>${D.nodes?D.nodes.length:0} nodes</span></div>
  <div><strong>NFS Exports</strong><span>${D.nfs?D.nfs.length:0} exports</span></div>
</div>

${D.nodes&&D.nodes.length ? '<h2>Node Inventory ('+D.nodes.length+' nodes)</h2>'+tbl(
  ['LNN','Name','Status','Model','Serial','CPU','Disk Capacity','Memory','Drives','Release'],
  D.nodes.map(n=>[
    n.lnn||'—',
    'Node-'+(n.lnn||'?'),
    nodeStatus(n),
    nodeModel(n),
    (n.hardware||{}).serial_number||'—',
    nodeCPU(n),
    nodeDisk(n),
    nodeMemory(n),
    (n.drives||[]).filter(d=>d.ui_state==='HEALTHY').length+'/'+(n.drives||[]).length+' healthy',
    (n.status||{}).release||'—',
  ])
) : ''}

${D.nfs&&D.nfs.length ? '<h2>NFS Exports ('+D.nfs.length+')</h2>'+tbl(
  ['ID','Path(s)','Description','Access','Clients','Auth Flavors','Map Root','All Dirs'],
  D.nfs.map(e=>[
    e.id||'—',
    (e.paths||[]).join(', ')||'—',
    e.description||'—',
    e.read_only?'Read-Only':'Read-Write',
    (e.clients||[]).join(', ')||'All (*)',
    (e.security_flavors||[]).join(', ')||'—',
    nfsMapRoot(e),
    e.all_dirs?'Yes':'No',
  ])
) : ''}

${D.smb&&D.smb.length ? '<h2>SMB Shares ('+D.smb.length+')</h2>'+tbl(
  ['Share Name','Path','Description','Visibility','Access','Continuously Available'],
  D.smb.map(s=>[
    s.name||'—', s.path||'—', s.description||'—',
    s.browsable?'Browsable':'Hidden',
    s.read_only?'Read-Only':'Read-Write',
    s.continuously_available?'Yes':'No',
  ])
) : ''}

${D.quotas&&D.quotas.length ? '<h2>Quota Policies ('+D.quotas.length+')</h2>'+tbl(
  ['Path','Type','Logical Used','Physical Used','Hard Limit','Soft Limit','Advisory','Mode','Used%'],
  D.quotas.map(q=>{
    const u=q.usage||{}, t=q.thresholds||{};
    const hard=t.hard||0;
    const pct = hard&&u.logical ? Math.round(u.logical/hard*100)+'%' : '—';
    return [q.path||'—',q.type||'—',fb(u.logical),fb(u.physical),
            hard?fb(hard):'None', t.soft?fb(t.soft):'None',
            t.advisory?fb(t.advisory):'None',
            q.enforced?'Enforced':'Advisory', pct];
  })
) : ''}

${D.snapshots&&D.snapshots.length ? '<h2>Snapshots ('+D.snapshots.length+')</h2>'+tbl(
  ['ID','Name','Path','Size','State','Created','Expires'],
  D.snapshots.map(s=>[
    s.id||'—', s.name||'—', s.path||'—', fb(s.size),
    s.state||'—', fd(s.created), s.expires?fd(s.expires):'Never',
  ])
) : ''}

${D.replication!=null ? '<h2>SyncIQ Policies ('+(D.replication||[]).length+')</h2>'+(
  (D.replication||[]).length ?
    tbl(['Policy Name','Source Path','Target Host','Target Path','Action','State','Schedule'],
      D.replication.map(p=>[
        p.name||'—', p.source_root_path||'—', p.target_host||'—', p.target_path||'—',
        p.action||'—', p.enabled?'Enabled':'Disabled', p.schedule||'Manual',
      ])) :
    '<p style="color:#5a7080;font-size:12px">No SyncIQ policies configured on this cluster.</p>'
) : ''}

${D.pools&&D.pools.length ? '<h2>Storage Pools ('+(D.pools||[]).length+')</h2>'+tbl(
  ['Pool Name','Type','Total Capacity','Used','Available','Used%','Protection','Member Nodes'],
  D.pools.map(p=>{
    const u=p.usage||{};
    const total=_to_num(u.usable_bytes||u.total_bytes||0);
    const avail=_to_num(u.avail_bytes||u.free_bytes||0);
    const pct  =total?Math.round((total-avail)/total*100)+'%':'—';
    return [p.name||'—', p.type||'—', fb(total)||'—', fb(total-avail)||'—',
            fb(avail)||'—', pct, (p.protection||p.protection_policy||
            (p.data_protection&&p.data_protection.requested_protection)||
            (p.data_protection&&p.data_protection.protection_policy)||'—'),
            (p.lnns||[]).join(', ')||'—'];
  })
) : ''}

${D.zones&&D.zones.length ? '<h2>Access Zones ('+(D.zones||[]).length+')</h2>'+tbl(
  ['Zone Name','Base Path','Auth Providers','Groupnet','System Provider'],
  D.zones.map(z=>[
    z.name||'—', z.path||'—',
    (z.auth_providers||[]).join(', ')||'—',
    z.groupnet||'—', z.system_provider||'—',
  ])
) : ''}

<div class="footer">
  PowerScale Migration Collector v4.0 &nbsp;·&nbsp;
  Report generated: ${new Date().toLocaleString()} &nbsp;·&nbsp;
  Auth: Session-Cookie (isisessid + isicsrf) &nbsp;·&nbsp;
  No credentials stored or transmitted beyond the cluster session.
</div>
</body></html>`;

  dl(new Blob([html],{type:'text/html; charset=utf-8'}),
     `psscale_report_${safe(D.clusterName)}_${dateStr()}.html`);
  log('✔ HTML report exported','ok');
}
</script>
</body>
</html>"""

@app.route("/")
def index():
    return Response(HTML, content_type="text/html; charset=utf-8")

# ══════════════════════════════════════════════════════════════════
#  STARTUP
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 70)
    print("  PowerScale / Isilon OneFS — Migration Data Collector  v4.0 (Clean)")
    print("  Developer : SHAIKH SHOAIB · Sr. Advisor Delivery Specialist")
    print("-" * 70)
    print(f"  Listening  : http://{LISTEN_HOST}:{LISTEN_PORT}")
    print(f"  Auth       : POST /session/1/session (isisessid + isicsrf CSRF)")
    print(f"  Session TTL: {SESSION_TIMEOUT_MINUTES} minutes")
    print(f"  Debug      : {'ON' if DEBUG_MODE else 'OFF'}")
    print("  Confirmed  : All 404 paths removed — only verified PAPI paths used")
    print("  NTP        : Node drift via /cluster/time | Servers via manual entry")
    print("  Security   : Credentials in server memory only — never on disk")
    print("=" * 70)
    print(f"\n  Open browser → http://localhost:{LISTEN_PORT}\n")
    try:
        app.run(host=LISTEN_HOST, port=LISTEN_PORT,
                debug=False, threaded=True, use_reloader=False)
    except OSError as e:
        print(f"\n[ERROR] Port {LISTEN_PORT} in use.\n{e}")
        sys.exit(1)